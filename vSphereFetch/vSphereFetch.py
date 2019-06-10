#!/usr/bin/env python3
#coding=utf-8
# VMware vSphere Python SDK
# Copyright (c) 2008-2013 VMware, Inc. All Rights Reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""
Python program for listing the vms on an ESX / vCenter host
"""

import os, time, atexit
from collections import OrderedDict

from pyVim import connect
from pyVmomi import vmodl
from pyVmomi import vim
from datetime import timedelta, datetime

#from tkinter.filedialog import *
#from tkinter import font
from tkinter import ttk, Tk, Frame, Label, Entry, Button, Text, Scrollbar, LEFT, RIGHT, END, Y, font

#import argparse
#import getpass

from openpyxl import Workbook
#from openpyxl.cell import get_column_letter
from openpyxl.writer.write_only import WriteOnlyCell
from openpyxl.styles import Alignment, Font

import requests
requests.packages.urllib3.disable_warnings()
import ssl
try:
    _create_unverified_https_context = ssl._create_unverified_context
except AttributeError:
    pass
else:
    ssl._create_default_https_context = _create_unverified_https_context
__author__ = "VMware, Inc."

maps = OrderedDict()
maps['datacenter'] = '所在平台'
maps['cluster'] = '所在集群'
maps['name'] = '虚拟机名称'
maps['describe'] = '描述'
maps['system'] = '操作系统'
maps['ip'] = 'IP地址'
maps['status'] = '状态'
maps['vmwaretools'] = '增强工具'
maps['template'] = '模版'
maps['snapshot'] = '系统快照'
maps['question'] = '报告'
maps['containertype'] = '容器类型'
maps['productinfo'] = '产品信息'
maps['faulttolerance'] = '容错策略'
maps['vmxpath'] = '数据文件'
maps['vdisks'] = '虚拟磁盘'
maps['vnics'] = '虚拟PCI'
maps['limits'] = '资源限额'
maps['reservations'] = '资源预留'
maps['vcnum'] = 'CPU个数'
maps['vcready'] = 'CPU就绪'
maps['vcusage'] = 'CPU使用率'
maps['vmtotal'] = '内存大小'
maps['vmshared'] = '内存共享'
maps['vmballoon'] = '内存回收'
maps['vmswapped'] = '内存交换'
maps['vmactive'] = '活动内存'
maps['vsio'] = '存储IO'
maps['vslatency'] = '存储延迟'
maps['vnusage'] = '网络使用率'
maps['hname'] = '宿主机名称'
maps['hcdetail'] = '宿主机CPU核数'
maps['hctype'] = '宿主机CPU类型'
maps['hcusage'] = '宿主机CPU使用率'
maps['hmusage'] = '宿主机内存使用率'

class Vmachine:
    datacenter = ''
    cluster = ''
    name = ''
    describe = ''
    system = ''
    ip = ''
    status = ''
    vmwaretools = ''
    template = ''
    snapshot = ''
    question = ''
    containertype = ''
    productinfo = ''
    faulttolerance = ''
    vmxpath = ''
    vdisks = ''
    vnics = ''
    limits = ''
    reservations = ''
    vcnum = ''
    vcready = ''
    vcusage = ''
    vmtotal = ''
    vmshared = ''
    vmballoon = ''
    vmswapped = ''
    vmactive = ''
    vsio = ''
    vslatency = ''
    vnusage = ''
    hname = ''
    hcdetail = ''
    hctype = ''
    hcusage = ''
    hmusage = ''
    
    def __init__(self, datacenter, cluster, hname, name):
        self.datacenter = datacenter
        self.cluster = cluster
        self.hname = hname
        self.name = name
        
    def dd(self, datacenter, cluster, hname, name, describe, system, ip, status, vmwaretools, template, snapshot, question, containertype, productinfo, faulttolerance, vmxpath, 
                vdisks, vnics, limits, reservations, vcnum, vcready, vcusage, vmtotal, vmshared, vmballoon, vmswapped, vmactive, vsio, vslatency, vnusage, hcdetail, hctype, hcusage, hmusage):
        self.datacenter = datacenter
        self.cluster = cluster
        self.hname = hname
        self.name = name
        self.describe = describe
        self.system = system
        self.ip = ip
        self.status = status
        self.vmwaretools = vmwaretools
        self.template = template
        self.snapshot = snapshot
        self.question = question
        self.containertype = containertype
        self.productinfo = productinfo
        self.faulttolerance = faulttolerance
        self.vmxpath = vmxpath
        self.vdisks = vdisks
        self.vnics = vnics
        self.limits = limits
        self.reservations = reservations
        self.vcnum = vcnum
        self.vcready = vcready
        self.vcusage = vcusage
        self.vmtotal = vmtotal
        self.vmshared = vmshared
        self.vmballoon = vmballoon
        self.vmswapped = vmswapped
        self.vmactive = vmactive
        self.vsio = vsio
        self.vslatency = vslatency
        self.vnusage = vnusage
        self.hcdetail = hcdetail
        self.hctype = hctype
        self.hcusage = hcusage
        self.hmusage = hmusage

    def getVar(self, data):
        if data == 'datacenter':
            return self.datacenter
        elif data == 'cluster':
            return self.cluster
        elif data == 'hname':
            return self.hname
        elif data == 'name':
            return self.name
        elif data == 'describe':
            return self.describe
        elif data == 'system':
            return self.system
        elif data == 'ip':
            return self.ip
        elif data == 'status':
            return self.status
        elif data == 'vmwaretools':
            return self.vmwaretools
        elif data == 'template':
            return self.template
        elif data == 'snapshot':
            return self.snapshot
        elif data == 'question':
            return self.question
        elif data == 'containertype':
            return self.containertype
        elif data == 'productinfo':
            return self.productinfo
        elif data == 'faulttolerance':
            return self.faulttolerance
        elif data == 'vmxpath':
            return self.vmxpath
        elif data == 'vdisks':
            return self.vdisks
        elif data == 'vnics':
            return self.vnics
        elif data == 'limits':
            return self.limits
        elif data == 'reservations':
            return self.reservations
        elif data == 'vcnum':
            return self.vcnum
        elif data == 'vcready':
            return self.vcready
        elif data == 'vcusage':
            return self.vcusage
        elif data == 'vmtotal':
            return self.vmtotal
        elif data == 'vmshared':
            return self.vmshared
        elif data == 'vmballoon':
            return self.vmballoon
        elif data == 'vmswapped':
            return self.vmswapped
        elif data == 'vmactive':
            return self.vmactive
        elif data == 'vsio':
            return self.vsio
        elif data == 'vslatency':
            return self.vslatency
        elif data == 'vnusage':
            return self.vnusage
        elif data == 'hcdetail':
            return self.hcdetail
        elif data == 'hctype':
            return self.hctype
        elif data == 'hcusage':
            return self.hcusage
        elif data == 'hmusage':
            return self.hmusage
        
    def __str__(self):
        return( 'datacenter        : ' + self.datacenter + '\n' +
               'cluster           : ' + self.cluster + '\n' + 
               'name              : ' + self.name + '\n' + 
               'describe          : ' + self.describe + '\n' +
               'system            : ' + self.system + '\n' +
               'ip                : ' + self.ip + '\n' +
               'status            : ' + self.status + '\n' +
               'vmwaretools       : ' + self.vmwaretools + '\n' +
               'template          : ' + self.template + '\n' +
               'snapshot          : ' + self.snapshot + '\n' +
               'question          : ' + self.question + '\n' +
               'containertype     : ' + self.containertype + '\n' +
               'productinfo       : ' + self.productinfo + '\n' +
               'faulttolerance    : ' + self.faulttolerance + '\n' +
               'vmxpath           : ' + self.vmxpath + '\n' +
               'vdisks            : ' + self.vdisks + '\n' +
               'vnics             : ' + self.vnics + '\n' +
               'limits            : ' + self.limits + '\n' +
               'reservations      : ' + self.reservations + '\n' +
               'vcnum             : ' + self.vcnum + '\n' +
               'vcready           : ' + self.vcready + '\n' +
               'vcusage           : ' + self.vcusage + '\n' +
               'vmtotal           : ' + self.vmtotal + '\n' +
               'vmshared          : ' + self.vmshared + '\n' +
               'vmballoon         : ' + self.vmballoon + '\n' +
               'vmswapped         : ' + self.vmswapped + '\n' +
               'vmactive          : ' + self.vmactive + '\n' +
               'vsio              : ' + self.vsio + '\n' +
               'vslatency         : ' + self.vslatency + '\n' +
               'vnusage           : ' + self.vnusage + '\n' +
               'hname             : ' + self.hname + '\n' + 
               'hcdetail          : ' + self.hcdetail + '\n' +
               'hctype            : ' + self.hctype + '\n' +
               'hcusage           : ' + self.hcusage + '\n' +
               'hmusage           : ' + self.hmusage + '\n')
        

#def build_arg_parser():
#    """
#    Builds a standard argument parser with arguments for talking to vCenter

#    -s service_host_name_or_ip
#    -o optional_port_number
#    -u required_user
#    -p optional_password

#    """
#    parser = argparse.ArgumentParser(description='Standard Arguments for talking to vCenter')

#    # because -h is reserved for 'help' we use -s for service
#    parser.add_argument('-s', '--host', required=True, action='store',help='vSphere service to connect to')

#    # because we want -p for password, we use -o for port
#    parser.add_argument('-o', '--port', type=int, default=443, action='store', help='Port to connect on')

#    parser.add_argument('-u', '--user', required=True, action='store', help='User name to use when connecting to host')

#    parser.add_argument('-p', '--password', required=False, action='store', help='Password to use when connecting to host')
#    return parser

#def prompt_for_password(args):
#    """
#    if no password is specified on the command line, prompt for it
#    """
#    if not args.password:
#        args.password = getpass.getpass(prompt='Enter password for host %s and user %s: ' %(args.host, args.user))
#    return args

#def get_args():
#    """
#    Supports the command-line arguments needed to form a connection to vSphere.
#    """
#    parser = build_arg_parser()

#    args = parser.parse_args()

#    return prompt_for_password(args)

#def getProperties(content, viewType, props, specType):
#    # Build a view and get basic properties for all Virtual Machines
#    objView = content.viewManager.CreateContainerView(content.rootFolder, viewType, True)
#    tSpec = vim.PropertyCollector.TraversalSpec(name='tSpecName', path='view', skip=False, type=vim.view.ContainerView)
#    pSpec = vim.PropertyCollector.PropertySpec(all=False, pathSet=props, type=specType)
#    oSpec = vim.PropertyCollector.ObjectSpec(obj=objView, selectSet=[tSpec], skip=False)
#    pfSpec = vim.PropertyCollector.FilterSpec(objectSet=[oSpec], propSet=[pSpec], reportMissingObjectsInResults=False)
#    retOptions = vim.PropertyCollector.RetrieveOptions()
#    totalProps = []
#    retProps = content.propertyCollector.RetrievePropertiesEx(specSet=[pfSpec], options=retOptions)
#    totalProps += retProps.objects
#    while retProps.token:
#        retProps = content.propertyCollector.ContinueRetrievePropertiesEx(token=retProps.token)
#        totalProps += retProps.objects
#    objView.Destroy()
#    # Turn the output in retProps into a usable dictionary of values
#    gpOutput = []
#    for eachProp in totalProps:
#        propDic = {}
#        for prop in eachProp.propSet:
#            propDic[prop.name] = prop.val
#        propDic['moref'] = eachProp.obj
#        gpOutput.append(propDic)
#    return gpOutput

def freshText(index, text, color):
    if index == 'end':
        logtext.insert(END, text, color)
    else:
        logtext.insert(0.0, text, color)
    logtext.see(END)
    logtext.update()

    


def statCheck(perf_dict, counter_name):
    counter_key = perf_dict[counter_name]
    return counter_key

def buildQuery(content, vchtime, counterId, instance, vm, interval):
    try:
        perfManager = content.perfManager
        metricId = vim.PerformanceManager.MetricId(counterId=counterId, instance=instance)
        startTime = vchtime - timedelta(minutes=(interval + 1))
        endTime = vchtime - timedelta(minutes=1)
        query = vim.PerformanceManager.QuerySpec(intervalId=20, entity=vm, metricId=[metricId], startTime=startTime, endTime=endTime)
        perfResults = perfManager.QueryPerf(querySpec=[query])
        if perfResults:
            return perfResults
        else:
            #print('ERROR: Performance results empty.  TIP: Check time drift on source and vCenter server')
            #print('Troubleshooting info:')
            #print('vCenter/host date and time: {}'.format(vchtime))
            #print('Start perf counter time   :  {}'.format(startTime))
            #print('End perf counter time     :  {}'.format(endTime))
            #print(query)
            return None
    except Exception as ex:
        return None

def getVmInfo(vmobj, vm, content, vchtime, interval, perf_dict):
    statInt = interval * 3  # There are 3 20s samples in each minute
    summary = vm.summary
    disk_list = []
    network_list = []

    # Convert limit and reservation values from -1 to None
    try:
        if vm.resourceConfig.cpuAllocation.limit == -1:
            vmcpulimit = "None"
        else:
            vmcpulimit = "{} Mhz".format(vm.resourceConfig.cpuAllocation.limit)
        if vm.resourceConfig.memoryAllocation.limit == -1:
            vmmemlimit = "None"
        else:
            vmmemlimit = "{} MB".format(vm.resourceConfig.cpuAllocation.limit)

        if vm.resourceConfig.cpuAllocation.reservation == 0:
            vmcpures = "None"
        else:
            vmcpures = "{} Mhz".format(vm.resourceConfig.cpuAllocation.reservation)
        if vm.resourceConfig.memoryAllocation.reservation == 0:
            vmmemres = "None"
        else:
            vmmemres = "{} MB".format(vm.resourceConfig.memoryAllocation.reservation)
        
        #hsmy = summary.runtime.host.summary
        vm_hardware = vm.config.hardware
        if vm_hardware is not None:
            for each_vm_hardware in vm_hardware.device:
                if (each_vm_hardware.key >= 2000) and (each_vm_hardware.key < 3000):
                    disk_list.append('{} | {:.1f}GB | Thin: {} | {}'.format(each_vm_hardware.deviceInfo.label, each_vm_hardware.capacityInKB/1024/1024, each_vm_hardware.backing.thinProvisioned, each_vm_hardware.backing.fileName))
                elif (each_vm_hardware.key >= 4000) and (each_vm_hardware.key < 5000):
                    network_list.append('{} | {} | {}'.format(each_vm_hardware.deviceInfo.label, each_vm_hardware.deviceInfo.summary, each_vm_hardware.macAddress))

        #CPU Ready Average
        #if vm.summary.runtime.powerState == 'poweredOn':
        statCpuReady = buildQuery(content, vchtime, (statCheck(perf_dict, 'cpu.ready.summation')), "", vm, interval)
        if statCpuReady is not None:
            cpuReady = (float(sum(statCpuReady[0].value[0].value)) / statInt)
        else:
            cpuReady = None
        #CPU Usage Average % - NOTE: values are type LONG so needs divided by 100 for percentage
        statCpuUsage = buildQuery(content, vchtime, (statCheck(perf_dict, 'cpu.usage.average')), "", vm, interval)
        if statCpuUsage is not None:
            cpuUsage = ((float(sum(statCpuUsage[0].value[0].value)) / statInt) / 100)
        else:
            cpuUsage = None
        #Memory Active Average MB
        statMemoryActive = buildQuery(content, vchtime, (statCheck(perf_dict, 'mem.active.average')), "", vm, interval)
        if statMemoryActive is not None:
            memoryActive = (float(sum(statMemoryActive[0].value[0].value) / 1024) / statInt)
        else:
            memoryActive = None
        #Memory Shared
        statMemoryShared = buildQuery(content, vchtime, (statCheck(perf_dict, 'mem.shared.average')), "", vm, interval)
        if statMemoryShared is not None:
            memoryShared = (float(sum(statMemoryShared[0].value[0].value) / 1024) / statInt)
        else:
            memoryShared = None
        #Memory Balloon
        statMemoryBalloon = buildQuery(content, vchtime, (statCheck(perf_dict, 'mem.vmmemctl.average')), "", vm, interval)
        if statMemoryBalloon is not None:
            memoryBalloon = (float(sum(statMemoryBalloon[0].value[0].value) / 1024) / statInt)
        else:
            memoryBalloon = None
        #Memory Swapped
        statMemorySwapped = buildQuery(content, vchtime, (statCheck(perf_dict, 'mem.swapped.average')), "", vm, interval)
        if statMemorySwapped is not None:
            memorySwapped = (float(sum(statMemorySwapped[0].value[0].value) / 1024) / statInt)
        else:
            memorySwapped = None
        #Datastore Average IO
        statDatastoreIoRead = buildQuery(content, vchtime, (statCheck(perf_dict, 'datastore.numberReadAveraged.average')), "*", vm, interval)
        if statDatastoreIoRead is not None:
            DatastoreIoRead = (float(sum(statDatastoreIoRead[0].value[0].value)) / statInt)
        else:
            DatastoreIoRead = None
        statDatastoreIoWrite = buildQuery(content, vchtime, (statCheck(perf_dict, 'datastore.numberWriteAveraged.average')), "*", vm, interval)
        if statDatastoreIoWrite is not None:
            DatastoreIoWrite = (float(sum(statDatastoreIoWrite[0].value[0].value)) / statInt)
        else:
            DatastoreIoWrite = None
        #Datastore Average Latency
        statDatastoreLatRead = buildQuery(content, vchtime, (statCheck(perf_dict, 'datastore.totalReadLatency.average')), "*", vm, interval)
        if statDatastoreLatRead is not None:
            DatastoreLatRead = (float(sum(statDatastoreLatRead[0].value[0].value)) / statInt)
        else:
            DatastoreLatRead = None
        statDatastoreLatWrite = buildQuery(content, vchtime, (statCheck(perf_dict, 'datastore.totalWriteLatency.average')), "*", vm, interval)
        if statDatastoreLatWrite is not None:
            DatastoreLatWrite = (float(sum(statDatastoreLatWrite[0].value[0].value)) / statInt)
        else:
            DatastoreLatWrite = None
        #Network usage (Tx/Rx)
        statNetworkTx = buildQuery(content, vchtime, (statCheck(perf_dict, 'net.transmitted.average')), "", vm, interval)
        if statNetworkTx is not None:
            networkTx = (float(sum(statNetworkTx[0].value[0].value) * 8 / 1024) / statInt)
        else:
            networkTx = None
        statNetworkRx = buildQuery(content, vchtime, (statCheck(perf_dict, 'net.received.average')), "", vm, interval)
        if statNetworkRx is not None:
            networkRx = (float(sum(statNetworkRx[0].value[0].value) * 8 / 1024) / statInt)
        else:
            networkRx = None

        
        #print('\nNOTE: Any VM statistics are averages of the last {} minutes\n'.format(statInt / 3))
        #print('Server Name                    :', summary.config.name)
        vmobj.name = str(summary.config.name)
        
        #print('Description                    :', summary.config.annotation)
        vmobj.describe = str(summary.config.annotation)
        
        #print('Guest                          :', summary.config.guestFullName)
        vmobj.system = str(summary.config.guestFullName)
        
        if summary.guest is not None:
            #print('IP Address                     :', summary.guest.ipAddress)
            #print('VMware Tools                   :', summary.guest.toolsStatus)
            vmobj.ip = str(summary.guest.ipAddress)
            vmobj.vmwaretools = str(summary.guest.toolsStatus)
        else:
            #print('IP Address                     : Unknown')
            #print('VMware Tools                   : Unknown')
            vmobj.ip = 'Unknown'
            vmobj.vmwaretools = 'Unknown'
        
        #print('Template                       : ', summary.config.template)
        vmobj.template = str(summary.config.template)
        	
        if vm.rootSnapshot:
            #print('Snapshot Status                : Snapshots present')
            vmobj.snapshot = 'Snapshots present'
        else:
            #print('Snapshot Status                : No Snapshots')
            vmobj.snapshot = 'No Snapshots'
        
        if summary.runtime.question is not None:
            #print('Question                       :', summary.runtime.question.text)
            vmobj.question = str(summary.runtime.question.text)
        else:
            #print('Question                       : No Questions')
            vmobj.question = 'No Questions'
        
        #print('Container Type                 :', summary.config.guestId)
        vmobj.containertype = str(summary.config.guestId)
        
        #print('Product Info                   :', summary.config.product)
        vmobj.productinfo = str(summary.config.product)
        
        #print('Fault Tolerance                :', summary.config.ftInfo)
        vmobj.faulttolerance = str(summary.config.ftInfo)
        
        #print('VM .vmx Path                   :', summary.config.vmPathName)
        vmobj.vmxpath = str(summary.config.vmPathName)
        
        if len(disk_list) > 0:
            #print('Virtual Disks                  :', disk_list[0])
            vmobj.vdisks = str(disk_list[0])
        else:
            #print('Virtual Disks                  : None')
            vmobj.vdisks = 'None'
        if len(disk_list) > 1:
            disk_list.pop(0)
            for each_disk in disk_list:
                #print('                                ', each_disk)
                vmobj.vdisks += ' * ' + str(each_disk)
        
        if len(network_list) > 0:
            #print('Virtual NIC(s)                 :', network_list[0])
            vmobj.vnics = network_list[0]
        else:
            #print('Virtual NIC(s)                 :None')
            vmobj.vnics = 'None'
        if len(network_list) > 1:
            network_list.pop(0)
            for each_vnic in network_list:
                #print('                                ', each_vnic)
                vmobj.vnics += ' * ' + str(each_vnic)
        
        #print('Status                         :', summary.runtime.powerState) 
        vmobj.status = str(summary.runtime.powerState)   
        
        #print('[VM] Limits                    : CPU: {}, Memory: {}'.format(vmcpulimit, vmmemlimit))
        vmobj.limits = 'CPU: {}, Memory: {}'.format(vmcpulimit, vmmemlimit)
        
        #print('[VM] Reservations              : CPU: {}, Memory: {}'.format(vmcpures, vmmemres))
        vmobj.reservations = 'CPU: {}, Memory: {}'.format(vmcpures, vmmemres)
        
        #print('[VM] Number of vCPUs           :', summary.config.numCpu)
        vmobj.vcnum = str(summary.config.numCpu)

        if cpuReady is None:
            #print('[VM] CPU Ready                 : Unknown')
            vmobj.vcready = 'Unknown'
        else:
            #print('[VM] CPU Ready                 : Average {:.1f} %, Maximum {:.1f} %'.format((cpuReady / 20000 * 100), ((float(max(statCpuReady[0].value[0].value)) / 20000 * 100))))
            vmobj.vcready = 'Average {:.1f} %, Maximum {:.1f} %'.format((cpuReady / 20000 * 100), ((float(max(statCpuReady[0].value[0].value)) / 20000 * 100)))
            
        if cpuUsage is None:
            #print('[VM] CPU (%)                   : Unknown')
            vmobj.vcusage = 'Unknown'
        else:
            #print('[VM] CPU (%)                   : {:.0f} %'.format(cpuUsage))
            vmobj.vcusage = '{:.0f} %'.format(cpuUsage)
            
        #print('[VM] Memory                    : {} MB ({:.1f} GB)'.format(summary.config.memorySizeMB, (float(summary.config.memorySizeMB) / 1024)))
        vmobj.vmtotal = '{} MB ({:.1f} GB)'.format(summary.config.memorySizeMB, (float(summary.config.memorySizeMB) / 1024))
        
        if memoryShared is None:
            #print('[VM] Memory Shared             : Unknown')
            vmobj.vmshared = 'Unknown'
        else:
            #print('[VM] Memory Shared             : {:.0f} %, {:.0f} MB'.format(((memoryShared / summary.config.memorySizeMB) * 100), memoryShared))
            vmobj.vmshared = '{:.0f} %, {:.0f} MB'.format(((memoryShared / summary.config.memorySizeMB) * 100), memoryShared)
            
        if memoryBalloon is None:
            #print('[VM] Memory Balloon            : Unknown')
            vmobj.vmballoon = 'Unknown'
        else:
            #print('[VM] Memory Balloon            : {:.0f} %, {:.0f} MB'.format(((memoryBalloon / summary.config.memorySizeMB) * 100), memoryBalloon))
            vmobj.vmballoon = '{:.0f} %, {:.0f} MB'.format(((memoryBalloon / summary.config.memorySizeMB) * 100), memoryBalloon)
            
        if memorySwapped is None:
            #print('[VM] Memory Swapped            : Unknown')
            vmobj.vmswapped = 'Unknown'
        else:
            #print('[VM] Memory Swapped            : {:.0f} %, {:.0f} MB'.format(((memorySwapped / summary.config.memorySizeMB) * 100), memorySwapped))
            vmobj.vmswapped = '{:.0f} %, {:.0f} MB'.format(((memorySwapped / summary.config.memorySizeMB) * 100), memorySwapped)
            
        if memoryActive is None:
            #print('[VM] Memory Active             : Unknown')
            vmobj.vmactive = 'Unknown'
        else:
            #print('[VM] Memory Active             : {:.0f} %, {:.0f} MB'.format(((memoryActive / summary.config.memorySizeMB) * 100), memoryActive))
            vmobj.vmactive = '{:.0f} %, {:.0f} MB'.format(((memoryActive / summary.config.memorySizeMB) * 100), memoryActive)
            
        if DatastoreIoRead is None or DatastoreIoWrite is None:
            #print('[VM] Datastore Average IO      : Read: Unknown, Write: Unknown')
            vmobj.vsio = 'Read: Unknown, Write: Unknown'
        else:
            #print('[VM] Datastore Average IO      : Read: {:.0f} IOPS, Write: {:.0f} IOPS'.format(DatastoreIoRead, DatastoreIoWrite))
            vmobj.vsio = 'Read: {:.0f} IOPS, Write: {:.0f} IOPS'.format(DatastoreIoRead, DatastoreIoWrite)
            
        if DatastoreLatRead is None or DatastoreLatWrite is None:
            #print('[VM] Datastore Average Latency : Read: Unknown, Write: Unknown')
            vmobj.vslatency = 'Read: Unknown, Write: Unknown'
        else:
            #print('[VM] Datastore Average Latency : Read: {:.0f} ms, Write: {:.0f} ms'.format(DatastoreLatRead, DatastoreLatWrite))
            vmobj.vslatency = 'Read: {:.0f} ms, Write: {:.0f} ms'.format(DatastoreLatRead, DatastoreLatWrite)
            
        if networkTx is None or networkRx is None:
            #print('[VM] Overall Network Usage     : Transmitted Unknown, Received Unknown')
            vmobj.vnusage = 'Transmitted Unknown, Received Unknown'
        else:
            #print('[VM] Overall Network Usage     : Transmitted {:.3f} Mbps, Received {:.3f} Mbps'.format(networkTx, networkRx))
            vmobj.vnusage = 'Transmitted {:.3f} Mbps, Received {:.3f} Mbps'.format(networkTx, networkRx)
            
        #print('[Host] Name                    : {}'.format(summary.runtime.host.name))
        #if hsmy is not None:
        #    print('[Host] CPU Detail              : Processor Sockets: {}, Cores per Socket {}'.format(hsmy.hardware.numCpuPkgs, (hsmy.hardware.numCpuCores / hsmy.hardware.numCpuPkgs)))
        #    print('[Host] CPU Type                : {}'.format(hsmy.hardware.cpuModel))
        #    print('[Host] CPU Usage               : Used: {} Mhz, Total: {} Mhz'.format(hsmy.quickStats.overallCpuUsage, (hsmy.hardware.cpuMhz * hsmy.hardware.numCpuCores)))
        #    print('[Host] Memory Usage            : Used: {:.0f} GB, Total: {:.0f} GB'.format((float(hsmy.quickStats.overallMemoryUsage) / 1024), (float(hsmy.hardware.memorySize) / 1024 / 1024 / 1024)))
        #else:
        #    print('[Host] CPU Detail              : Unknown')
        #    print('[Host] CPU Type                : Unknown')
        #    print('[Host] CPU Usage               : Used: Unknown, Total: Unknown')
        #    print('[Host] Memory Usage            : Used: Unknown, Total: Unknown')
        #print('\n')
    except Exception as ex:
        print(ex)

def collect(chost, cuser, cpasswd, cport):
    """
    Simple command-line program for listing the virtual machines on a system.
    """
    #args = get_args()
    vms = []
    #sys.stdout = open('stdout.txt', 'w')

    try:
        freshText('end', '\n正在连接到vCenter ' + chost + ' ...', 'orange')
        #service_instance = connect.SmartConnect(host=args.host, user=args.user, pwd=args.password, port=int(args.port))
        service_instance = connect.SmartConnect(host=chost, user=cuser, pwd=cpasswd, port=int(cport))
        atexit.register(connect.Disconnect, service_instance)
        content = service_instance.RetrieveContent()
        vchtime = service_instance.CurrentTime()

        # Get all the performance counters
        perf_dict = {}
        perfList = content.perfManager.perfCounter
        for counter in perfList:
            counter_full = "{}.{}.{}".format(counter.groupInfo.key, counter.nameInfo.key, counter.rollupType)
            perf_dict[counter_full] = counter.key

        #retProps = getProperties(content, [vim.VirtualMachine], ['name', 'runtime.powerState'], vim.VirtualMachine)

        container = content.rootFolder  # starting point to look into
        #if container is not None and isinstance(container, vim.Folder):
            #dc_moref = container.CreateDatacenter(name=dcname)
        #viewType = [vim.VirtualMachine]  # object types to look for
        #viewType = [vim.ClusterComputeResource] #get cluster
        viewType = [vim.Datacenter]   #get Datacenter
        recursive = True  # whether we should look into it recursively
        containerView = content.viewManager.CreateContainerView(container, viewType, recursive)

        datacenters = containerView.view
        freshText('end', '[成功]', 'green')
        freshText('end', '\n开始收集虚拟机信息 ...\n', 'orange')
        for datacenter in datacenters:
            #print(datacenter.name)
            for cluster in datacenter.hostFolder.childEntity:
                #print('\t' + cluster.name)
                if hasattr(cluster, 'host'):
                    for host in cluster.host:
                        #print('\t\t' + host.name)
                        hcdetail = 'Processor Sockets: {}, Cores per Socket {}'.format(host.summary.hardware.numCpuPkgs, (host.summary.hardware.numCpuCores / host.summary.hardware.numCpuPkgs))
                        hctype = '{}'.format(host.summary.hardware.cpuModel)
                        hcusage = 'Used: {} Mhz, Total: {} Mhz'.format(host.summary.quickStats.overallCpuUsage, (host.summary.hardware.cpuMhz * host.summary.hardware.numCpuCores))
                        hmusage = 'Used: {:.0f} GB, Total: {:.0f} GB'.format((float(host.summary.quickStats.overallMemoryUsage) / 1024), (float(host.summary.hardware.memorySize) / 1024 / 1024 / 1024))
                        if hasattr(host, 'vm'):
                            for vm in host.vm:
                                vmobj = Vmachine(datacenter.name, cluster.name, host.name, vm.name)
                                vmobj.hcdetail = hcdetail
                                vmobj.hctype = hctype
                                vmobj.hcusage = hcusage
                                vmobj.hmusage = hmusage
                                #print(vmobj)
                                #print('\t\t\t' + vm.name)
                                #print('\n' + cluster.name + ':' + vm.name)
                                #if vm.summary.runtime.powerState != 'poweredOn':
                                    #print(vm.summary.runtime.host.name)
                                freshText('end', '\n正在收集虚拟机 ', 'orange')
                                freshText('end', vmobj.datacenter + ' <-> ' + vmobj.cluster + ' <-> ' + vmobj.name, 'deeppink')
                                freshText('end', ' 信息 ...', 'orange')
                                try:
                                    getVmInfo(vmobj, vm, content, vchtime, 15, perf_dict)
                                    vms.append(vmobj)
                                    #logtext.insert(END, '********************************************************************************\n')
                                    #logtext.insert(END, str(vmobj))
                                    freshText('end', '[完成]\n', 'green')
                                except Exception as ex:
                                    freshText('end', '[失败]\n', 'red')
                                    continue
                                #vms.append(Vmachine(datacenter.name, cluster.name, host.name, vm.name, ''))
                #if hasattr(cluster, 'summary'):
                #    print('\t\t' + str(cluster.summary))   
        #viewType = [vim.HostSystem]
        #recursive = True  # whether we should look into it recursively
        #containerView = content.viewManager.CreateContainerView(
        #    container, viewType, recursive)
        #hosts = containerView.view
        #for host in hosts:
        #    print(host.name)
            #if hasattr(host, 'summary'):
            #    print('\t' + str(host.summary))
            
        #info = open('vms.txt', 'w')
        #info.truncate()
        #for vm in vms:
        #    info.write('#######################################################################################################\n')
        #    info.write(str(vm))
        #info.close()
        
        #print('Total: ' + str(len(vms)))
        return vms
    except Exception as ex:
        freshText('end', '\n发生错误: ' + str(ex) + '\n', 'red')
        #print("Caught vmodl fault : " + str(ex))
        return None

    return None

def export(vms, type):
    ctime = time.strftime('%Y.%m.%d.%H.%M.%S', time.localtime(time.time()))
    filename = '虚拟机统计表' + ctime + '.xlsx'
    filepath = os.path.join(os.getcwd(), filename)
    wb = Workbook(write_only=False)
    ws = wb.create_sheet(title='虚拟机统计表')
    titlelist = []
    for colname in maps.values():
        titlecell = WriteOnlyCell(ws, value=colname)
        titlecell.font = Font(bold=True)
        titlecell.alignment = Alignment(horizontal='center', vertical='center')
        titlelist.append(titlecell)
    ws.append(titlelist)
    for vm in vms:
        celllist = []
        for coltitle in maps.keys():
            celllist.append(vm.getVar(coltitle))
        ws.append(celllist)
    wb.remove_sheet(wb.active)
    wb.save(filepath)
    return filepath

def run():
    host = hostentry.get()
    port = portentry.get()
    user = userentry.get()
    passwd = passwdentry.get()
    logtext.delete(0.0, END)
    logtext.see(END)
    logtext.update()
    if host is not None and host.strip() != '' and port is not None and port.strip() != '' and user is not None and user.strip() != '' and passwd is not None and passwd.strip() != '':
        try:
            vms = collect(host, user, passwd, int(port))
            if vms is not None:
                freshText('end', '\n信息收集完成!', 'green')
                if len(vms) > 0:
                    freshText('end', '\n正在保存信息 ...', 'orange')
                    try:
                        filepath = export(vms, 'xlsx')
                        freshText('end', '\n信息保存完成: ' + filepath, 'green')
                    except Exception as ex:
                        freshText('end', '\n信息保存失败: ' + str(ex), 'red')
                        return
                else:
                    freshText('end', '\n没有发现虚拟机!', 'pink')
            else:
                freshText('end', '\n信息收集失败!', 'red')
        except Exception as ex:
            freshText('end', '\n发生错误: ' + str(ex), 'red')
    else:
        freshText('end', '\nvCenter相关信息填写不全!', 'red')

def clr():
    hostentry.delete(0, END)
    portentry.delete(0, END)
    userentry.delete(0, END)
    passwdentry.delete(0, END)
    logtext.delete(0.0, END)
    #hostentry.insert(END, 'test')
    #logtext.insert(END, 'aaa')
    
win = Tk()
win.wm_title('vSphere抓取器')
tfrm = Frame(win)
tfrm.pack()
cfrm = Frame(win)
cfrm.pack()

tifrm = Frame(tfrm)
tifrm.grid(row=0, column=0, columnspan=1, sticky='w')
trfrm = Frame(tfrm)
trfrm.grid(row=0, column=1, columnspan=1, sticky='e')

#win.geometry('400*100')
hostlabel = Label(tifrm, text='vCenter地址 :', fg='olivedrab', font=("微软雅黑", 9, "bold"))
hostentry = Entry(tifrm, fg='chocolate', font=("微软雅黑", 9, "bold"))
portlabel = Label(tifrm, text='vCenter端口 :', fg='olivedrab', font=("微软雅黑", 9, "bold"))
portentry = Entry(tifrm, fg='chocolate', font=("微软雅黑", 9, "bold"))
userlabel = Label(tifrm, text='帐            号 :', fg='olivedrab', font=("微软雅黑", 9, "bold"))
userentry  = Entry(tifrm, fg='chocolate', font=("微软雅黑", 9, "bold"))
passwdlabel = Label(tifrm, text='密            码 :', fg='olivedrab', font=("微软雅黑", 9, "bold"))
passwdentry = Entry(tifrm,fg='chocolate', show='*', font=("微软雅黑", 9, "bold"))

hostlabel.grid(row=0, column=0, padx=5, pady=3)
hostentry.grid(row=0, column=1, padx=5, pady=3)
portlabel.grid(row=0, column=2, padx=5, pady=3)
portentry.grid(row=0, column=3, padx=5, pady=3)
userlabel.grid(row=1, column=0, padx=5, pady=3)
userentry.grid(row=1, column=1, padx=5, pady=3)
passwdlabel.grid(row=1, column=2, padx=5, pady=3)
passwdentry.grid(row=1, column=3, padx=5, pady=3)

rbtn = Button(trfrm, text='抓取', command=run, fg='blueviolet', font=("微软雅黑", 9, "bold"))
cbtn = Button(trfrm, text='清空', command=clr, fg='blueviolet', font=("微软雅黑", 9, "bold"))
rbtn.grid(row=0, column=0, sticky='e')
cbtn.grid(row=1, column=0, sticky='e')

efrm = Frame(cfrm)
efrm.grid(row=1, column=0, padx=3, pady=3)
#logtext = Text(win, height=15, wrap='none')  #设置滚动条-不换行  
logtext = Text(efrm,)
vscrollbar = Scrollbar(efrm, command=logtext.yview)  #文本框-竖向滚动条  
#hscrollbar = Scrollbar(efrm, command=logtext.xview)  #文本框-横向滚动条  
logtext.config(yscrollcommand=vscrollbar.set)
#vscrollbar.config(command=logtext.yview)  
#hscrollbar.config(command=logtext.xview) 
logtext.pack(side=LEFT, fill=Y)
vscrollbar.pack(side=RIGHT, fill=Y)
logtext.bind('<KeyPress>', lambda e:'break')
logtext.tag_config('red',foreground = 'red')
logtext.tag_config('blue',foreground = 'blue')
logtext.tag_config('green',foreground = 'green')
logtext.tag_config('black',foreground = 'black')
logtext.tag_config('orange',foreground = 'orange')
logtext.tag_config('gold',foreground = 'gold')
logtext.tag_config('cyan',foreground = 'cyan')
logtext.tag_config('purple',foreground = 'purple')
logtext.tag_config('pink',foreground = 'pink')
logtext.tag_config('deeppink',foreground = 'deeppink')
logtext.tag_config('gray',foreground = 'gray')

hostentry.insert(END, '110.50.1.15')
portentry.insert(END, '443')
userentry.insert(END, 'administrator')
passwdentry.insert(END, 'password')
logtext.insert(END, '\n\n\n\n\n\n')
logtext.insert(END, '                                ******************\n', 'blue')
logtext.insert(END, '                                *    欢迎使用    *\n', 'blue')
logtext.insert(END, '                                ******************\n', 'blue')
logtext.insert(END, '\n\n\n\n\n\n\n\n\n\n\n\n')
logtext.insert(END, '  Programmed By XSHRIM        Mail: chenpan@hrbb.com.cn        Tel: 17080056739\n', 'purple')
logtext.insert(END, ' -------------------------------------------------------------------------------', 'gray')
    
win.mainloop()

#if __name__ == "__main__":
#    #args = get_args()
#    main('', '', '', '')
#using like "python test.py -s 110.50.1.15 -u administrator -p password"