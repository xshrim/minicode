import random

import pymysql
from openpyxl import *
from openpyxl.utils import get_column_letter

config = {
    'host': '110.1.1.30',
    'port': 3306,
    'user': 'root',
    'passwd': 'Admin#233',
    'db': 'datacenter',
    'charset': 'utf8'
}


# create a connection
# conf: parameters of the connection
def connDB(conf):
    conn = pymysql.connect(**conf)
    return conn


# return query result
# conn: connection name;
# sql: query statement;
# args: query parameters;
# fetchnum: fetch number.
def exeQuery(conn, tsql, args, fetchnum):
    res = None
    try:
        with conn.cursor(pymysql.cursors.DictCursor) as cur:
            cur.execute(tsql, args)
            if fetchnum is None or fetchnum == -1:
                res = cur.fetchall()
            else:
                res = cur.fetchmany(fetchnum)
        conn.commit()
    finally:
        conn.close()
    return res


def exeUpdate(conn, tsql, args):
    res = None
    try:
        with conn.cursor() as cur:
            res = cur.execute(tsql, args)
        conn.commit()
    finally:
        conn.close()
    return res


def connClose(conn, cur):
    cur.close()
    conn.close()


# export data from Excel file
# filepath: path of target Excel file
# sheet: the sheet of Excel file need to be exported
# beginrow: begin exporting from this row
# begincol: begin exporting from this column
def expData(filepath, sheet, beginrow, begincol, pt, area, vcenter, vcaccount):
    title = None
    datalist = []
    coldict = {}
    #rowdict = {}
    colnum = 0
    rownum = beginrow - 1
    wb = load_workbook(filepath)
    ws = wb[sheet]

    # get the real count of columns
    for colindex in range(1, ws.max_column):
        #print(chr(ord('A') + colindex) + str(beginrow-1))
        if ws[get_column_letter(colindex) + str(beginrow-1)].value is not None:
            colnum += 1
            #coldict[get_column_letter(colindex)] = ws[get_column_letter(colindex) + str(beginrow-1)].value.strip()
        else:
            break

    # get the real count of rows
    for rowindex in range(beginrow, ws.max_row + 1):
        if ws['A' + str(rowindex)].value is not None or ws['B' + str(rowindex)].value is not None or ws['C' + str(rowindex)].value is not None or ws['D' + str(rowindex)].value is not None:
            rownum += 1
        else:
            break

    # export data
    for row in range(beginrow, rownum + 1):
        rowdata = {}
        for col in range(1, ws.max_column):
            #colname = coldict[chr(ord(begincol) + col - (ord(begincol) - ord('A')))]
            colname = ws[get_column_letter(col) + str(beginrow-1)].value
            colvalue = ws[get_column_letter(col) + str(row)].value
            if colvalue is not None:
                colvalue = str(colvalue).strip()
            if colvalue is None or colvalue == '' or colvalue == 'None':
                colvalue = '-'
            colvalue = colvalue.replace('：', ':').strip()

            #rowdata['principal'] = '-'
            #rowdata['uptime'] = '-'
            if colname == '所在平台':
                rowdata['platform'] = pt + ':' + colvalue
            elif colname == '所在集群':
                rowdata['cluster'] = colvalue
            elif colname == '虚拟机名称':
                print(colvalue)
                rowdata['name'] = colvalue
            elif colname == '描述':
                rowdata['notes'] = colvalue
                rowdata['describe'] = colvalue
                desclist = colvalue.split('\n')
                for desc in desclist:
                    if '系统:' in desc:
                        rowdata['describe'] = desc.split(':')[1]
                    if '应用管理员:' in desc:
                        rowdata['principal'] = desc.split(':')[1]
                    if '创建日期:' in desc:
                        rowdata['uptime'] = desc.split(':')[1]
                    if '说明:' in desc:
                        rowdata['notes'] = desc.split(':')[1]
            elif colname == '操作系统':
                rowdata['system'] = colvalue
            elif colname == 'IP地址':
                rowdata['ip'] = colvalue
            elif colname == '状态':
                if colvalue == 'poweredOn':
                    rowdata['status'] = '在线'
                else:
                    rowdata['status'] = '关机'
            elif colname == '增强工具':
                if colvalue == 'toolsOk':
                    rowdata['vmwaretools'] = '开启'
                else:
                    rowdata['vmwaretools'] = '关闭'
            elif colname == '系统快照':
                if colvalue == 'Snapshots present':
                    rowdata['snapshot'] = '开启'
                else:
                    rowdata['snapshot'] = '关闭'
            elif colname == '虚拟磁盘':
                    rowdata['hdisk'] = colvalue
            elif colname == '虚拟PCI':
                rowdata['pci'] = colvalue
            elif colname == 'CPU个数':
                rowdata['cpu'] = colvalue + ' vCPU'
            elif colname == '内存大小':
                rowdata['memory'] = colvalue
            elif colname == '宿主机名称':
                rowdata['phyname'] = colvalue
        if rowdata['system'] is not None:
            if 'windows' in str(rowdata['system']).lower():
                rowdata['loginmethod'] = 'RDP'
            else:
                rowdata['loginmethod'] = 'SSH'
        else:
            rowdata['loginmethod'] = '-'
        rowdata['project'] = '-'
        if '村行' in pt:
            rowdata['domain'] = '村行'
        else:
            rowdata['domain'] = '哈行'
        rowdata['area'] = area
        rowdata['hostname'] = '-'
        rowdata['account'] = '-'

        rowdata['backuper'] = '-'

        rowdata['monitored'] = '未知'
        rowdata['inntp'] = '未知'
        rowdata['ansible'] = '未知'
        rowdata['vcenter'] = vcenter
        rowdata['vcaccount'] = vcaccount
        rowdata['belong'] = '系统组'
        rowdata['implementor'] = '张鹏/13351786818'
        datalist.append(rowdata)
    return datalist


def impData(data):
    res = ''
    connection = connDB(config)
    for item in data:
        colname = ''
        colloc = ''
        colvalue = []
        for col in item:
            if item[col] is not None:
                colname += '`' + col + '`, '
                colloc += '%s, '
                colvalue.append(item[col])
        colname = colname[:-2]
        colloc = colloc[:-2]
        tsql = 'insert ignore into vdevice(' + colname + ') values(' + colloc + ')'
        try:
            with connection.cursor() as cur:
                res = cur.execute(tsql, colvalue)
        except Exception as ex:
            print(str(ex))
            #print(tsql)
    connection.commit()
    connection.close()
    print(res)
        #sq = "select roomlocation, racknum from device"
        #sql = 'insert into device(roomlocation, racknum, ulocation) values(%s, %s, %s)'
        #params = ('国裕三楼', 'RA-01', '9-12')
        #result = exeQuery(connection, sq, None, None)
        #columns = ','.join(item.keys())
        #values = item.values()
        #for ckey in item:
            #print(ckey)

'''
vm6data = expData('VM6虚拟化虚拟机统计表2016.12.19.09.15.15.xlsx', '虚拟机统计表', 2, 'A', 'VM6虚拟化', '生产', '35.1.7.101', 'administrator@hrbbvm6.com/Admin@123')
for item in vm6data:
    if item['describe'] is not None:
        descs = item['describe'].replace('：', ':').strip()
        desclist = descs.split('\n')

        for desc in desclist:
            if '系统:' in desc:
                print(desc.split(':')[1])
            if '应用管理员:' in desc:
                print(desc.split(':')[1])
            if '创建日期:' in desc:
                print(desc.split(':')[1])
            if '说明:' in desc:
                print(desc.split(':')[1])
    print('================================')
'''


impData(expData('OA虚拟化.xlsx', '虚拟机统计表', 2, 'A', 'OA虚拟化', 'OA', '110.50.1.15', 'administrator/password'))
impData(expData('VM6虚拟化.xlsx', '虚拟机统计表', 2, 'A', 'VM6虚拟化', '生产', '35.1.7.101', 'administrator@hrbbvm6.com/Admin@123'))
#impData(expData('村行虚拟化.xlsx', '虚拟机统计表', 2, 'A', '村行虚拟化', '生产', '116.1.7.6', 'administrator/password@01'))
impData(expData('临时虚拟化.xlsx', '虚拟机统计表', 2, 'A', '临时虚拟化', 'OA', '168.1.10.43', 'administrator/password@01'))
impData(expData('事中事后国库虚拟化.xlsx', '虚拟机统计表', 2, 'A', '事中事后国库虚拟化', '生产', '118.63.1.53', 'administrator/P@ssw0rd'))
impData(expData('网银虚拟化.xlsx', '虚拟机统计表', 2, 'A', '网银虚拟化', '生产', '35.8.0.12', 'administrator/root@123'))
impData(expData('一期虚拟化.xlsx', '虚拟机统计表', 2, 'A', '一期虚拟化', '生产', '50.1.1.180', 'administrator/vmware123'))
impData(expData('二期虚拟化.xlsx', '虚拟机统计表', 2, 'A', '二期虚拟化', '生产', '50.1.2.20', 'administrator/vmware123'))
impData(expData('三期虚拟化.xlsx', '虚拟机统计表', 2, 'A', '三期虚拟化', '生产', '35.1.7.26', 'administrator/P@ssw0rd'))
impData(expData('容灾一期虚拟化.xlsx', '虚拟机统计表', 2, 'A', '容灾一期虚拟化', '生产', '118.63.254.26', 'administrator/password'))
impData(expData('容灾二期虚拟化.xlsx', '虚拟机统计表', 2, 'A', '容灾二期虚拟化', '生产', '118.63.254.214', 'administrator/vmware123'))
impData(expData('准生产虚拟化.xlsx', '虚拟机统计表', 2, 'A', '准生产虚拟化', '准生产', '35.1.7.103', 'administrator@zscvm6.com/Root@123'))



# connection.commit()
# connClose(connection, cursor)
