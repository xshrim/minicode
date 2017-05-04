#!/usr/bin/env python3
# coding=utf-8

# programmed by ChenPan

import time
from tkinter.filedialog import *
from tkinter import font
from tkinter import ttk
from tkinter.messagebox import *
# from openpyxl import *
from openpyxl import load_workbook, Workbook
from openpyxl import styles
from openpyxl.writer.excel import ExcelWriter
from openpyxl.styles import Font, Border, Alignment, Side


def clock():
    endline['text'] = '欢迎使用'


def check(dic, flag):
    try:
        if flag == 'l':
            if dic['lserverTitle'] is None or dic['lserverTitle'].strip() == '' or dic['lserverInfo'] is None or dic['lserverInfo'].strip() == '' or dic['lswitchTitle'] is None or dic['lswitchTitle'].strip() == '' or dic['lswitchInfo'] is None or dic['lswitchInfo'].strip() == '':
                return False
            else:
                return True
        elif flag == 'm':
            if dic['mserverTitle'] is None or dic['mserverTitle'].strip() == '' or dic['mserverTitle'].strip() == '设备:' or dic['mserverInfo'] is None or dic['mserverInfo'].strip() == ''or dic['mserverInfo'].strip() == 'IP:':
                return False
            else:
                return True
        elif flag == 'e':
            if dic['eserverTitle'] is None or dic['eserverTitle'].strip() == '' or dic['eserverTitle'].strip() == '设备:' or dic['eserverInfo'] is None or dic['eserverInfo'].strip() == ''or dic['eserverInfo'].strip() == 'IP:':
                return False
            else:
                return True
    except Exception as err:
        return False
        print(err)


def cmp(dica, dicb, flag):
    try:
        if flag == 'l':
            if dica['lserverTitle'] == dicb['lserverTitle'] and dica['lserverInfo'] == dicb['lserverInfo'] and dica['lswitchTitle'] == dicb['lswitchTitle'] and dica['lswitchInfo'] == dicb['lswitchInfo']:
                return True
            else:
                return False
        elif flag == 'm':
            if dica['mserverTitle'] == dicb['mserverTitle'] and dica['mserverInfo'] == dicb['mserverInfo']:
                return True
            else:
                return False
        elif flag == 'e':
            if dica['eserverTitle'] == dicb['eserverTitle'] and dica['eserverInfo'] == dicb['eserverInfo']:
                return True
            else:
                return False
    except Exception as err:
        return False
        print(err)


def getFile(flag):
    fname = ''
    try:
        options = {}
        options['defaultextension'] = '.xlsx'
        options['filetypes'] = [('xlsx', '.xlsx')]
        # options['parent'] =
        if flag == 'l':
            options['title'] = '网络--导入xlsx文件'
        elif flag == 'm':
            options['title'] = '主机--导入xlsx文件'
        elif flag == 'e':
            options['title'] = '电源--导入xlsx文件'
        elif flag == 'g':
            options['title'] = '全局--导入xlsx文件'
        openfile = askopenfile(**options)
        fname = openfile.name
    except Exception as err:
        print(err)
    return fname


def getItem(flag):
    buff = []
    try:
        if flag == 'l':
            for children in ltree.get_children():
                litem = ltree.item(children)['values']
                buff.insert(0, {'lserverTitle': litem[0], 'lserverInfo': litem[2], 'lswitchTitle': litem[1], 'lswitchInfo': litem[3]})
        elif flag == 'm':
            for children in mtree.get_children():
                mitem = mtree.item(children)['values']
                buff.insert(0, {'mserverTitle': mitem[0], 'mserverInfo': mitem[1]})
        elif flag == 'e':
            for children in etree.get_children():
                eitem = etree.item(children)['values']
                buff.insert(0, {'eserverTitle': eitem[0], 'eserverInfo': eitem[1]})
    except Exception as err:
        print(err)
    return buff


def getMaxRow(ws, column):
    maxrow = 0
    for col in ws.columns:
        for cell in col:
            if cell.column == column and cell.value is not None:
                maxrow = max(cell.row, maxrow)
    return maxrow


def OnSelectL(event):
    try:
        for item in ltree.selection():
            litem = ltree.item(item)['values']
            lserverTitle_Entry.delete(0, END)
            lserverTitle_Entry.insert(0, litem[0])
            lserverInfo_Entry.delete(0, END)
            lserverInfo_Entry.insert(0, litem[2])
            lswitchTitle_Entry.delete(0, END)
            lswitchTitle_Entry.insert(0, litem[1])
            lswitchInfo_Entry.delete(0, END)
            lswitchInfo_Entry.insert(0, litem[3])
            break
    except Exception as err:
        print(err)


def OnSelectM(event):
    try:
        for item in mtree.selection():
            mitem = mtree.item(item)['values']
            mserverTitle_Entry.delete(0, END)
            mserverTitle_Entry.insert(0, mitem[0])
            mserverInfo_Entry.delete(0, END)
            mserverInfo_Entry.insert(0, mitem[1])
            break
    except Exception as err:
        print(err)


def OnSelectE(event):
    try:
        for item in etree.selection():
            eitem = etree.item(item)['values']
            eserverTitle_Entry.delete(0, END)
            eserverTitle_Entry.insert(0, eitem[0])
            eserverInfo_Entry.delete(0, END)
            eserverInfo_Entry.insert(0, eitem[1])
            break
    except Exception as err:
        print(err)


def ladd():
    try:
        flag = 0
        llabel = {'lserverTitle': lserverTitle_Entry.get(), 'lserverInfo': lserverInfo_Entry.get(), 'lswitchTitle': lswitchTitle_Entry.get(), 'lswitchInfo': lswitchInfo_Entry.get()}

        if not check(llabel, 'l'):
            flag = 1
            showinfo(title='Error', message='线标填写不规范!')
        else:
            item = getItem('l')
            if len(item) != 0:
                for label in item:
                    # if dic == label:
                    if cmp(label, llabel, 'l'):
                        flag = 1
                        showinfo(title='Error', message='该线标已经录入!')
                        break
            if flag == 0:
                lechoLabel['text'] = int(lechoLabel['text']) + 1
                ltree.insert('', 0, values=(llabel['lserverTitle'], llabel['lswitchTitle'], llabel['lserverInfo'], llabel['lswitchInfo']))
    except Exception as err:
        print(err)


def limp():
    try:
        duplicate = 0
        illegal = 0
        impcount = 0

        targetSheet = networksheet

        fname = getFile('l')
        if fname == '':
            print('读取文件出错!')
            return
        elif fname[fname.rfind('.'):] != '.xlsx':
            showinfo(title='Error', message='不支持的文件格式!')
            return

        wb = load_workbook(filename=fname)
        wss = wb.get_sheet_names()
        ws = None
        for sheet in wb.get_sheet_names():
            if sheet == targetSheet:
                ws = wb.get_sheet_by_name(sheet)
                break
        if ws is None and len(wss) > 1:
            ws = wb.get_sheet_by_name(wss[1])
        elif len(wss) == 1:
            ws = wb.get_sheet_by_name(wss[0])
        if ws is not None:
            for i in range(lnetworkrowbegin, ws.max_row + 1):
                flag = 0
                lserverTitle = ws.cell(lnetworkcola + str(i)).value
                lswitchTitle = ws.cell(lnetworkcolb + str(i)).value
                lserverInfo = ws.cell(lnetworkcolc + str(i)).valueis
                lswitchInfo = ws.cell(lnetworkcold + str(i)).value
                if lserverTitle is None or lswitchTitle is None or lserverInfo is None or lswitchInfo is None:
                    if (lserverTitle is None and lswitchTitle is None and lserverInfo is None and lswitchInfo is None) is False:
                        illegal = illegal + 1
                    llabel = None
                else:
                    llabel = {'lserverTitle': lserverTitle, 'lserverInfo': lserverInfo, 'lswitchTitle': lswitchTitle, 'lswitchInfo': lswitchInfo}
                for label in getItem('l'):
                    if cmp(label, llabel, 'l'):
                        flag = 1
                        duplicate = duplicate + 1
                        break
                if flag == 0 and llabel is not None:
                    impcount = impcount + 1
                    lechoLabel['text'] = int(lechoLabel['text']) + 1
                    ltree.insert('', 0, values=(llabel['lserverTitle'], llabel['lswitchTitle'], llabel['lserverInfo'], llabel['lswitchInfo']))
            if duplicate > 0 or illegal > 0:
                showinfo(title='Info', message=str(duplicate) + '条重复数据/' + str(illegal) + '条非法数据未导入！')
        else:
            showinfo(title='Error', message='没有找到合适的表格!')
    except Exception as err:
        print(err)


def ledit():
    try:
        for selecteditem in ltree.selection():
            flag = 0
            index = 0
            index = ltree.index(selecteditem)

            llabel = {'lserverTitle': lserverTitle_Entry.get(), 'lserverInfo': lserverInfo_Entry.get(), 'lswitchTitle': lswitchTitle_Entry.get(), 'lswitchInfo': lswitchInfo_Entry.get()}

            if not check(llabel, 'l'):
                flag = 1
                showinfo(title='Error', message='线标填写不规范!')
            else:
                item = getItem('l')
                if len(item) != 0:
                    for label in item:
                        # if dic == label:
                        if cmp(label, llabel, 'l'):
                            flag = 1
                            showinfo(title='Error', message='该线标已经录入!')
                            break
                if flag == 0:
                    ltree.delete(selecteditem)
                    ltree.insert('', index, values=(llabel['lserverTitle'], llabel['lswitchTitle'], llabel['lserverInfo'], llabel['lswitchInfo']))
            break
    except Exception as err:
        print(err)


def ldelete():
    try:
        for item in ltree.selection():
            ltree.delete(item)
        lechoLabel['text'] = len(ltree.get_children())
        return 0
    except Exception as err:
        print(err)


def lclear():
    try:
        for item in ltree.get_children():
            ltree.delete(item)
        lechoLabel['text'] = 0
    except Exception as err:
        print(err)


def madd():
    try:
        flag = 0

        mserverTitle = mserverTitle_Entry.get()
        mserverInfo = mserverInfo_Entry.get()
        if mserverTitle[0:3] != '设备:' and mserverTitle[0:3] != '设备：' and mserverTitle[0:3].upper() != 'DEVICE:' and mserverTitle[0:3].upper() != 'DEVICE：':
            mserverTitle = '设备:' + mserverTitle
        if mserverInfo[0:3].upper() != 'IP:' and mserverInfo[0:3].upper() != 'IP：':
            mserverInfo = 'IP:' + mserverInfo

        mlabel = {'mserverTitle': mserverTitle, 'mserverInfo': mserverInfo}

        if not check(mlabel, 'm'):
            flag = 1
            showinfo(title='Error', message='标签填写不规范!')
        else:
            item = getItem('m')
            if len(item) != 0:
                for label in item:
                    # if dic == label:
                    if cmp(label, mlabel, 'm'):
                        flag = 1
                        showinfo(title='Error', message='该标签已经录入!')
                        break
            if flag == 0:
                mechoLabel['text'] = int(mechoLabel['text']) + 1
                mtree.insert('', 0, values=(mlabel['mserverTitle'], mlabel['mserverInfo']))
    except Exception as err:
        print(err)


def mimp():
    try:
        duplicate = 0
        illegal = 0
        impcount = 0

        targetSheet = hostsheet

        fname = getFile('m')
        if fname == '':
            print('读取文件出错!')
            return
        elif fname[fname.rfind('.'):] != '.xlsx':
            showinfo(title='Error', message='不支持的文件格式!')
            return

        wb = load_workbook(filename=fname)
        wss = wb.get_sheet_names()
        ws = None
        for sheet in wb.get_sheet_names():
            if sheet == targetSheet:
                ws = wb.get_sheet_by_name(sheet)
                break
        if ws is None and len(wss) > 2:
            ws = wb.get_sheet_by_name(wss[2])
        elif len(wss) == 1:
            ws = wb.get_sheet_by_name(wss[0])
        if ws is not None:
            for i in range(mhostrowbegin, ws.max_row + 1):
                flag = 0
                mserverTitle = ws.cell(mhostcola + str(i)).value
                mserverInfo = ws.cell(mhostcolb + str(i)).value
                if mserverTitle is None or mserverInfo is None:
                    if (mserverTitle is None and mserverInfo is None) is False:
                        illegal = illegal + 1
                    mlabel = None
                else:
                    mlabel = {'mserverTitle': mserverTitle, 'mserverInfo': mserverInfo}
                for label in getItem('m'):
                    if cmp(label, mlabel, 'm'):
                        flag = 1
                        duplicate = duplicate + 1
                        break
                if flag == 0 and mlabel is not None:
                    impcount = impcount + 1
                    mechoLabel['text'] = int(mechoLabel['text']) + 1
                    mtree.insert('', 0, values=(mlabel['mserverTitle'], mlabel['mserverInfo']))
            if duplicate > 0 or illegal > 0:
                showinfo(title='Info', message=str(duplicate) + '条重复数据/' + str(illegal) + '条非法数据未导入！')
        else:
            showinfo(title='Error', message='没有找到合适的表格!')
    except Exception as err:
        print(err)


def medit():
    try:
        for selecteditem in mtree.selection():
            flag = 0
            index = 0
            index = mtree.index(selecteditem)

            mserverTitle = mserverTitle_Entry.get()
            mserverInfo = mserverInfo_Entry.get()
            if mserverTitle[0:3] != '设备:' and mserverTitle[0:3] != '设备：' and mserverTitle[0:3].upper() != 'DEVICE:' and mserverTitle[0:3].upper() != 'DEVICE：':
                mserverTitle = '设备:' + mserverTitle
            if mserverInfo[0:3].upper() != 'IP:' and mserverInfo[0:3].upper() != 'IP：':
                mserverInfo = 'IP:' + mserverInfo

            mlabel = {'mserverTitle': mserverTitle, 'mserverInfo': mserverInfo}

            if not check(mlabel, 'm'):
                flag = 1
                showinfo(title='Error', message='标签填写不规范!')
            else:
                item = getItem('m')
                if len(item) != 0:
                    for label in item:
                        # if dic == label:
                        if cmp(label, mlabel, 'm'):
                            flag = 1
                            showinfo(title='Error', message='该标签已经录入!')
                            break
                if flag == 0:
                    mtree.delete(selecteditem)
                    mtree.insert('', index, values=(mlabel['mserverTitle'], mlabel['mserverInfo']))
            break
    except Exception as err:
        print(err)


def mdelete():
    try:
        for item in mtree.selection():
            mtree.delete(item)
        mechoLabel['text'] = len(mtree.get_children())
    except Exception as err:
        print(err)


def mclear():
    try:
        for item in mtree.get_children():
            mtree.delete(item)
        mechoLabel['text'] = 0
    # mtree.delete(mtree.get_children())
    except Exception as err:
        print(err)


def eadd():
    try:
        flag = 0

        eserverTitle = eserverTitle_Entry.get()
        eserverInfo = eserverInfo_Entry.get()

        elabel = {'eserverTitle': eserverTitle, 'eserverInfo': eserverInfo}

        if not check(elabel, 'e'):
            flag = 1
            showinfo(title='Error', message='标签填写不规范!')
        else:
            item = getItem('e')
            if len(item) != 0:
                for label in item:
                    # if dic == label:
                    if cmp(label, elabel, 'e'):
                        flag = 1
                        showinfo(title='Error', message='该标签已经录入!')
                        break
            if flag == 0:
                eechoLabel['text'] = int(eechoLabel['text']) + 1
                etree.insert('', 0, values=(elabel['eserverTitle'], elabel['eserverInfo']))
    except Exception as err:
        print(err)


def eimp():
    try:
        duplicate = 0
        illegal = 0
        impcount = 0

        targetSheet = powersheet

        fname = getFile('e')
        if fname == '':
            print('读取文件出错!')
            return
        elif fname[fname.rfind('.'):] != '.xlsx':
            showinfo(title='Error', message='不支持的文件格式!')
            return

        wb = load_workbook(filename=fname)
        wss = wb.get_sheet_names()
        ws = None
        for sheet in wb.get_sheet_names():
            if sheet == targetSheet:
                ws = wb.get_sheet_by_name(sheet)
                break
        if ws is None and len(wss) > 3:
            ws = wb.get_sheet_by_name(wss[3])
        elif len(wss) == 1:
            ws = wb.get_sheet_by_name(wss[0])
        if ws is not None:
            for i in range(epowerrowbegin, ws.max_row + 1):
                flag = 0
                eserverTitle = ws.cell(epowercola + str(i)).value
                eserverInfo = ws.cell(epowercolb + str(i)).value
                if eserverTitle is None or eserverInfo is None:
                    if (eserverTitle is None and eserverInfo is None) is False:
                        illegal = illegal + 1
                    elabel = None
                else:
                    elabel = {'eserverTitle': eserverTitle, 'eserverInfo': eserverInfo}
                for label in getItem('e'):
                    if cmp(label, elabel, 'e'):
                        flag = 1
                        duplicate = duplicate + 1
                        break
                if flag == 0 and elabel is not None:
                    impcount = impcount + 1
                    eechoLabel['text'] = int(eechoLabel['text']) + 1
                    etree.insert('', 0, values=(elabel['eserverTitle'], elabel['eserverInfo']))
            if duplicate > 0 or illegal > 0:
                showinfo(title='Info', message=str(duplicate) + '条重复数据/' + str(illegal) + '条非法数据未导入！')
        else:
            showinfo(title='Error', message='没有找到合适的表格!')
    except Exception as err:
        print(err)


def eedit():
    try:
        for selecteditem in etree.selection():
            flag = 0
            index = 0
            index = etree.index(selecteditem)

            eserverTitle = eserverTitle_Entry.get()
            eserverInfo = eserverInfo_Entry.get()

            elabel = {'eserverTitle': eserverTitle, 'eserverInfo': eserverInfo}

            if not check(elabel, 'e'):
                flag = 1
                showinfo(title='Error', message='标签填写不规范!')
            else:
                item = getItem('e')
                if len(item) != 0:
                    for label in item:
                        # if dic == label:
                        if cmp(label, elabel, 'e'):
                            flag = 1
                            showinfo(title='Error', message='该标签已经录入!')
                            break
                if flag == 0:
                    etree.delete(selecteditem)
                    etree.insert('', index, values=(elabel['eserverTitle'], elabel['eserverInfo']))
            break
    except Exception as err:
        print(err)


def edelete():
    try:
        for item in etree.selection():
            etree.delete(item)
        eechoLabel['text'] = len(etree.get_children())
    except Exception as err:
        print(err)


def eclear():
    try:
        for item in etree.get_children():
            etree.delete(item)
        eechoLabel['text'] = 0
        # mtree.delete(mtree.get_children())
    except Exception as err:
        print(err)


def imp():
    try:
        lduplicate = 0
        lillegal = 0
        limpcount = 0
        mduplicate = 0
        millegal = 0
        mimpcount = 0
        eduplicate = 0
        eillegal = 0
        eimpcount = 0

        targetSheet = devicesheet

        # lbuff = []
        # mbuff = []
        # fileopt={}
        # fileopt['defaultextension'] = '.xlsx'
        # fileopt['filetypes'] = [('Excel Files','.xlsx')]
        fname = getFile('g')
        if fname == '':
            print('读取文件出错!')
            return
        elif fname[fname.rfind('.'):] != '.xlsx':
            showinfo(title='Error', message='不支持的文件格式!')
            return

        wb = load_workbook(filename=fname)
        wss = wb.get_sheet_names()
        ws = None
        for sheet in wb.get_sheet_names():
            if sheet == targetSheet:
                ws = wb.get_sheet_by_name(sheet)
                break
        if ws is None and len(wss) > 0:
            ws = wb.get_sheet_by_name(wss[0])
        if ws is not None:
            # lmaxrow = ws.max_row
            # mmaxrow = ws.max_row
            # emaxrow = ws.max_row
            # for k in range(lmaxrow, 3, -1):
                # if k > ldevicerowbegin and ws.cell(ldevicecola+str(k)).value is None and ws.cell(ldevicecolb+str(k)).value is None and ws.cell(ldevicecola+str(k-1)).value is None and ws.cell(ldevicecolb+str(k-1)).value is None :
                    # lmaxrow = lmaxrow - 1
                # if k > mdevicerowbegin and ws.cell(mdevicecol+str(k)).value is None and ws.cell(mdevicecol+str(k-1)).value is None:
                    # mmaxrow = mmaxrow - 1
                # if k > edevicerowbegin and ws.cell(edevicecola+str(k)).value is None:
                    # emaxrow = emaxrow - 1
            lmaxrow = max(getMaxRow(ws, ldevicecola), getMaxRow(ws, ldevicecolb))
            mmaxrow = getMaxRow(ws, mdevicecol)
            emaxrow = max(getMaxRow(ws, edevicecola), getMaxRow(ws, edevicecolb))

            for i in range(ldevicerowbegin, lmaxrow + 1, 2):
                lflag = 0
                lserverTitle = ws.cell(ldevicecola + str(i)).value
                lswitchTitle = ws.cell(ldevicecolb + str(i)).value
                lserverInfo = ws.cell(ldevicecola + str(i + 1)).value
                lswitchInfo = ws.cell(ldevicecolb + str(i + 1)).value
                if lserverTitle is None or lswitchTitle is None or lserverInfo is None or lswitchInfo is None:
                    if (lserverTitle is None and lswitchTitle is None and lserverInfo is None and lswitchInfo is None) is False:
                        lillegal = lillegal + 1
                    llabel = None
                else:
                    llabel = {'lserverTitle': lserverTitle, 'lserverInfo': lserverInfo, 'lswitchTitle': lswitchTitle, 'lswitchInfo': lswitchInfo}
                for label in getItem('l'):
                    if cmp(label, llabel, 'l'):
                        lflag = 1
                        lduplicate = lduplicate + 1
                        break
                if lflag == 0 and llabel is not None:
                    limpcount = limpcount + 1
                    lechoLabel['text'] = int(lechoLabel['text']) + 1
                    ltree.insert('', 0, values=(llabel['lserverTitle'], llabel['lswitchTitle'], llabel['lserverInfo'], llabel['lswitchInfo'],))

            for i in range(mdevicerowbegin, mmaxrow + 1, 2):
                mflag = 0
                mserverTitle = ws.cell(mdevicecol + str(i)).value
                mserverInfo = ws.cell(mdevicecol + str(i + 1)).value

                if mserverTitle is None or mserverInfo is None:
                    if (mserverTitle is None and mserverInfo is None) is False:
                        millegal = millegal + 1
                    mlabel = None
                else:
                    mlabel = {'mserverTitle': mserverTitle, 'mserverInfo': mserverInfo}
                for label in getItem('m'):
                    if cmp(label, mlabel, 'm'):
                        mflag = 1
                        mduplicate = mduplicate + 1
                        break
                if mflag == 0 and mlabel is not None:
                    mimpcount = mimpcount + 1
                    mechoLabel['text'] = int(mechoLabel['text']) + 1
                    mtree.insert('', 0, values=(mlabel['mserverTitle'], mlabel['mserverInfo']))

            for i in range(edevicerowbegin, emaxrow + 1):
                eflag = 0
                eserverTitle = ws.cell(edevicecola + str(i)).value
                eserverInfo = ws.cell(edevicecolb + str(i)).value
                if eserverTitle is None or eserverInfo is None:
                    if (eserverTitle is None and eserverInfo is None) is False:
                        eillegal = eillegal + 1
                    elabel = None
                else:
                    elabel = {'eserverTitle': eserverTitle, 'eserverInfo': eserverInfo}
                for label in getItem('e'):
                    if cmp(label, elabel, 'e'):
                        eflag = 1
                        eduplicate = eduplicate + 1
                        break
                if eflag == 0 and elabel is not None:
                    eimpcount = eimpcount + 1
                    eechoLabel['text'] = int(eechoLabel['text']) + 1
                    etree.insert('', 0, values=(elabel['eserverTitle'], elabel['eserverInfo']))

            if lduplicate + mduplicate + eduplicate > 0 or lillegal + millegal + eillegal > 0:
                showinfo(title='Info', message=str(lduplicate + mduplicate + eduplicate) + '条重复数据/' + str(lillegal + millegal + eillegal) + '条非法数据未导入！')

        else:
            showinfo(title='Error', message='没有找到合适的表格!')
    except Exception as err:
        print(err)


def clear():
    lclear()
    mclear()
    eclear()


# def clear():
#    global lbuff, mbuff
#    lbuff = []
#    mbuff = []
#    ltree.delete(ltree.get_children())
#    mtree.delete(mtree.get_children())
#    lechoLabel['text'] = 0
#    mechoLabel['text'] = 0
#    return 0


def generate():
    try:
        proj = ''

        lcoldim = defaultcoldim
        mcoldim = defaultcoldim
        ecoldim = defaultcoldim

        fontTitle = Font(name='Arial', size=10, bold=True)
        fontInfo = Font(name='Arial', size=10, bold=False)
        fontHead = Font(name='Arial', size=10, bold=True, color='FF0000')
        fontTitleHead = Font(name='Arial', size=10, bold=True, color='4B0082')
        borderThin = Border(left=Side(border_style=styles.borders.BORDER_THIN, color='FF000000'),
                            right=Side(border_style=styles.borders.BORDER_THIN, color='FF000000'),
                            top=Side(border_style=styles.borders.BORDER_THIN, color='FF000000'),
                            bottom=Side(border_style=styles.borders.BORDER_THIN, color='FF000000'),)
        alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=False, shrink_to_fit=False, indent=0)

        wb = Workbook()
        wb.create_sheet(devicesheet, 0)
        wb.create_sheet(networksheet, 1)
        wb.create_sheet(hostsheet, 2)
        wb.create_sheet(powersheet, 3)

        if wb.worksheets[4] is not None:
            wb.remove_sheet(wb.worksheets[4])
        ew = ExcelWriter(workbook=wb)

        wsa = wb.worksheets[0]
        # wsa.title = '设备标签表'
        wsb = wb.worksheets[2]
        wsc = wb.worksheets[3]
        wsd = wb.worksheets[1]
        # wsb.active = True

        wsalcount = ldevicerowbegin
        wsamcount = mdevicerowbegin
        wsaecount = edevicerowbegin
        wsbmcount = mhostrowbegin
        wscecount = epowerrowbegin
        wsdlcount = lnetworkrowbegin

        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            wsa.column_dimensions[col].width = defaultcoldim
            wsb.column_dimensions[col].width = defaultcoldim
            wsc.column_dimensions[col].width = defaultcoldim
            wsd.column_dimensions[col].width = defaultcoldim
        if ldevicerowbegin >= 2:
            ldevicetitle = ldevicerowbegin - 2
        if mdevicerowbegin >= 2:
            mdevicetitle = mdevicerowbegin - 2
        if edevicerowbegin >= 2:
            edevicetitle = edevicerowbegin - 2
        # ldevicetitlelocation = chr((ord(ldevicecola)+ord(ldevicecolb))//2)+str(ldevicetitle)
        ldevicetitlecells = ldevicecola + str(ldevicetitle) + ':' + ldevicecolb + str(ldevicetitle)
        mdevicetitlecells = mdevicecol + str(mdevicetitle)
        # edevicetitlelocation = chr((ord(edevicecola)+ord(edevicecolb))//2)+str(mdevicetitle)
        edevicetitlecells = edevicecola + str(edevicetitle) + ':' + edevicecolb + str(edevicetitle)
        wsa.merge_cells(ldevicetitlecells)
        wsa[ldevicecola + str(ldevicetitle)] = '网络标签'
        wsa[ldevicecola + str(ldevicetitle)].font = fontTitleHead
        wsa[ldevicecola + str(ldevicetitle)].alignment = alignment
        wsa[mdevicetitlecells] = '主机标签'
        wsa[mdevicetitlecells].font = fontTitleHead
        wsa[mdevicetitlecells].alignment = alignment
        wsa.merge_cells(edevicetitlecells)
        wsa[edevicecola + str(edevicetitle)] = '电源标签'
        wsa[edevicecola + str(edevicetitle)].font = fontTitleHead
        wsa[edevicecola + str(edevicetitle)].alignment = alignment
        wsa[ldevicecola + str(wsalcount - 1)] = '源'
        wsa[ldevicecola + str(wsalcount - 1)].font = fontHead
        wsa[ldevicecolb + str(wsalcount - 1)] = '目的'
        wsa[ldevicecolb + str(wsalcount - 1)].font = fontHead
        wsa[mdevicecol + str(wsamcount - 1)] = '源'
        wsa[mdevicecol + str(wsamcount - 1)].font = fontHead
        wsa[edevicecola + str(wsaecount - 1)] = '源'
        wsa[edevicecola + str(wsaecount - 1)].font = fontHead
        wsa[edevicecolb + str(wsaecount - 1)] = '目的'
        wsa[edevicecolb + str(wsaecount - 1)].font = fontHead

        wsb[mhostcola + str(wsbmcount - 1)] = 'C1'
        wsb[mhostcola + str(wsbmcount - 1)].font = fontHead
        wsb[mhostcolb + str(wsbmcount - 1)] = 'C2'
        wsb[mhostcolb + str(wsbmcount - 1)].font = fontHead

        wsc[epowercola + str(wscecount - 1)] = 'P1'
        wsc[epowercola + str(wscecount - 1)].font = fontHead
        wsc[epowercolb + str(wscecount - 1)] = 'P2'
        wsc[epowercolb + str(wscecount - 1)].font = fontHead

        wsd[lnetworkcola + str(wsdlcount - 1)] = 'F1'
        wsd[lnetworkcola + str(wsdlcount - 1)].font = fontHead
        wsd[lnetworkcolb + str(wsdlcount - 1)] = 'F2'
        wsd[lnetworkcolb + str(wsdlcount - 1)].font = fontHead
        wsd[lnetworkcolc + str(wsdlcount - 1)] = 'F3'
        wsd[lnetworkcolc + str(wsdlcount - 1)].font = fontHead
        wsd[lnetworkcold + str(wsdlcount - 1)] = 'F4'
        wsd[lnetworkcold + str(wsdlcount - 1)].font = fontHead

        for llabel in getItem('l'):
            tmplTitle = llabel['lserverTitle'].replace('（', '(')
            if '(' in tmplTitle:
                tmplTitle = tmplTitle[:tmplTitle.index('(')]
            if proj == '':
                proj = tmplTitle
            elif proj != tmplTitle and proj != '多系统':
                proj = '多系统'

            wsa[ldevicecola + str(wsalcount)] = llabel['lserverTitle']
            # wsa[ldevicecola+str(wsalcount)].alignment = alignment
            wsa[ldevicecolb + str(wsalcount)] = llabel['lswitchTitle']
            wsa[ldevicecola + str(wsalcount + 1)] = llabel['lserverInfo']
            wsa[ldevicecolb + str(wsalcount + 1)] = llabel['lswitchInfo']
            for i in range(ord(ldevicecola), ord(ldevicecolb) + 1):
                wsa[chr(i) + str(wsalcount)].font = fontTitle
                wsa[chr(i) + str(wsalcount)].border = borderThin
                wsa[chr(i) + str(wsalcount + 1)].font = fontInfo
                wsa[chr(i) + str(wsalcount + 1)].border = borderThin
            wsalcount = wsalcount + 2

            wsd[lnetworkcola + str(wsdlcount)] = llabel['lserverTitle']
            wsd[lnetworkcolb + str(wsdlcount)] = llabel['lswitchTitle']
            wsd[lnetworkcolc + str(wsdlcount)] = llabel['lserverInfo']
            wsd[lnetworkcold + str(wsdlcount)] = llabel['lswitchInfo']
            for i in range(ord(lnetworkcola), ord(lnetworkcold) + 1):
                wsd[chr(i) + str(wsdlcount)].border = borderThin
                if i <= (ord(lnetworkcola) + ord(lnetworkcold)) // 2:
                    wsd[chr(i) + str(wsdlcount)].font = fontTitle
                else:
                    wsd[chr(i) + str(wsdlcount)].font = fontInfo
            wsdlcount = wsdlcount + 1

            if llabel['lserverTitle'] is None:
                llabel['lserverTitle'] = ''
            if llabel['lswitchTitle'] is None:
                llabel['lwitchTitle'] = ''
            if llabel['lserverInfo'] is None:
                llabel['lserverInfo'] = ''
            if llabel['lswitchInfo'] is None:
                llabel['lswitchInfo'] = ''
            lcoldim = max(lcoldim, len(llabel['lserverTitle']) * 1.7, len(llabel['lswitchTitle']) * 1.7, len(llabel['lserverInfo']) * 1.3, len(llabel['lswitchInfo']) * 1.3)

        for mlabel in getItem('m'):
            tmpmTitle = mlabel['mserverTitle'].replace('：', ':')
            tmpmTitle = tmpmTitle.replace('设备:', '')
            if '_' in tmpmTitle:
                tmpmTitle = tmpmTitle[:tmpmTitle.index('_')]
            if proj == '':
                proj = tmpmTitle
            elif proj != tmpmTitle and proj != '多系统':
                tmproj = proj
                if len(proj) > 2:
                    tmproj = proj[:len(proj) - 2]
                if tmproj not in tmpmTitle:
                    proj = '多系统'

            tmpmTitle = mlabel['mserverTitle'].replace('：', ':')
            wsa[mdevicecol + str(wsamcount)] = mlabel['mserverTitle']
            wsa[mdevicecol + str(wsamcount + 1)] = mlabel['mserverInfo']
            for i in range(ord(mdevicecol), ord(mdevicecol) + 1):
                wsa[chr(i) + str(wsamcount)].font = fontTitle
                wsa[chr(i) + str(wsamcount)].border = borderThin
                wsa[chr(i) + str(wsamcount + 1)].font = fontInfo
                wsa[chr(i) + str(wsamcount + 1)].border = borderThin
            wsamcount = wsamcount + 2

            wsb[mhostcola + str(wsbmcount)] = mlabel['mserverTitle']
            wsb[mhostcolb + str(wsbmcount)] = mlabel['mserverInfo']
            for i in range(ord(mhostcola), ord(mhostcolb) + 1):
                wsb[chr(i) + str(wsbmcount)].border = borderThin
                if i % 2 == 1:
                    wsb[chr(i) + str(wsbmcount)].font = fontTitle
                else:
                    wsb[chr(i) + str(wsbmcount)].font = fontInfo
            wsbmcount = wsbmcount + 1

            if mlabel['mserverTitle'] is None:
                mlabel['mserverTitle'] = ''
            if mlabel['mserverInfo'] is None:
                mlabel['mserverInfo'] = ''
            mcoldim = max(mcoldim, len(mlabel['mserverTitle']) * 1.7, len(mlabel['mserverInfo']) * 1.3)

        for elabel in getItem('e'):
            # tmpeInfo = elabel['eserverInfo'].replace('：', ':')
            # if ':' in tmpeInfo:
                # tmpeInfo = tmpeInfo[:tmpeInfo.index(':')]
            # if proj == '':
                # proj = tmpeInfo
            # elif proj !=tmpeInfo and proj != '多系统':
                # tmproj = proj
                # if len(proj) > 2:
                    # tmproj = proj[:len(proj)-2]
                # if tmproj not in tmpeInfo:
                    # proj = '多系统'

            wsa[edevicecola + str(wsaecount)] = elabel['eserverTitle']
            wsa[edevicecola + str(wsaecount)].font = fontTitle
            wsa[edevicecolb + str(wsaecount)] = elabel['eserverInfo']
            wsa[edevicecolb + str(wsaecount)].font = fontInfo
            for i in range(ord(edevicecola), ord(edevicecolb) + 1):
                wsa[chr(i) + str(wsaecount)].border = borderThin
            wsaecount = wsaecount + 1

            wsc[epowercola + str(wscecount)] = elabel['eserverTitle']
            wsc[epowercolb + str(wscecount)] = elabel['eserverInfo']
            for i in range(ord(epowercola), ord(epowercolb) + 1):
                wsc[chr(i) + str(wscecount)].border = borderThin
                if i % 2 == 1:
                    wsc[chr(i) + str(wscecount)].font = fontTitle
                else:
                    wsc[chr(i) + str(wscecount)].font = fontInfo
            wscecount = wscecount + 1

            if elabel['eserverTitle'] is None:
                elabel['eserverTitle'] = ''
            if elabel['eserverInfo'] is None:
                elabel['eserverInfo'] = ''
            ecoldim = max(ecoldim, len(elabel['eserverTitle']) * 1.7, len(elabel['eserverInfo']) * 1.3)
        tproj = project_Entry.get()
        if tproj.strip() == '' or tproj == '项目名称':
            fname = '\\' + proj + '标签表_' + time.strftime('%Y-%m-%d') + '.xlsx'
        else:
            fname = '\\' + tproj + '.xlsx'
        destfile = askdirectory() + fname

        wsa.column_dimensions[ldevicecola].width = lcoldim
        wsa.column_dimensions[ldevicecolb].width = lcoldim
        wsa.column_dimensions[mdevicecol].width = mcoldim
        wsa.column_dimensions[edevicecola].width = ecoldim
        wsa.column_dimensions[edevicecolb].width = ecoldim
        wsb.column_dimensions[mhostcola].width = mcoldim
        wsb.column_dimensions[mhostcolb].width = mcoldim
        wsc.column_dimensions[epowercola].width = ecoldim
        wsc.column_dimensions[epowercolb].width = ecoldim
        wsd.column_dimensions[lnetworkcola].width = lcoldim
        wsd.column_dimensions[lnetworkcolb].width = lcoldim
        wsd.column_dimensions[lnetworkcolc].width = lcoldim
        wsd.column_dimensions[lnetworkcold].width = lcoldim

        wb.save(filename=destfile)
        endline['text'] = '成功生产Excel文件：' + destfile.replace('/', '\\')
        root.after(3000, clock)
    except Exception as err:
        print(err)


def about():
    showinfo('关于', '哈尔滨银行标签生成器 V1.0 \n\n \tBy 陈盼 \n \tTel: 17080056739 \n \tEmail: chenpan@hrbb.com.cn')


ldevicecola = 'B'
ldevicecolb = 'D'
ldevicerowbegin = 4
lnetworkcola = 'A'
lnetworkcolb = 'B'
lnetworkcolc = 'C'
lnetworkcold = 'D'
lnetworkrowbegin = 2
mdevicecol = 'F'
mdevicerowbegin = 4
mhostcola = 'A'
mhostcolb = 'B'
mhostrowbegin = 2
edevicecola = 'H'
edevicecolb = 'J'
edevicerowbegin = 4
epowerrowbegin = 2
epowercola = 'A'
epowercolb = 'B'
devicesheet = '设备标签表'
hostsheet = '主机标签表'
networksheet = '网络标签表'
powersheet = '电源标签表'
ldevicetitle = 0
mdevicetitle = 0
edevicetitle = 0
defaultcoldim = 10

root = Tk()
lserverTitleText = StringVar()
lserverTitleText.set('服务器线标头')

lserverInfoText = StringVar()
lserverInfoText.set('服务器线标值')

lswitchTitleText = StringVar()
lswitchTitleText.set('交换机线标头')

lswitchInfoText = StringVar()
lswitchInfoText.set('交换机线标值')

mserverTitleText = StringVar()
mserverTitleText.set('服务器标签头')

mserverInfoText = StringVar()
mserverInfoText.set('服务器标签值')

eserverTitleText = StringVar()
eserverTitleText.set('电源标签头')

eserverInfoText = StringVar()
eserverInfoText.set('电源标签值')

projectText = StringVar()
projectText.set('项目名称')

# serverTitle_Label = Label(root, text='服务器线标头:').grid(row=0, column=0)
# switchTitle_Label = Label(root, text='交换机线标头:').grid(row=1, column=0)

bft = font.Font(size=10, weight=font.BOLD)
tft = font.Font(size=10,)
ltree = ttk.Treeview(root, selectmode='extended', columns=('col1', 'col2', 'col3', 'col4'))
ltree.column('#0', width=10, anchor='w')
ltree.column('col1', width=250, anchor='center')
ltree.column('col2', width=250, anchor='center')
ltree.column('col3', width=250, anchor='center')
ltree.column('col4', width=250, anchor='center')
# tree.heading('#0', text='')
ltree.heading('col1', text='服务器线标头', anchor='center')
ltree.heading('col2', text='交换机线标头', anchor='center')
ltree.heading('col3', text='服务器线标值', anchor='center')
ltree.heading('col4', text='交换机线标值', anchor='center')
lysb = ttk.Scrollbar(root, orient='vertical', command=ltree.yview)
lxsb = ttk.Scrollbar(root, orient='horizontal', command=ltree.xview)
ltree.configure(yscroll=lysb.set, xscroll=lxsb.set)
# ltree.heading('#0', text='Path', anchor='w')
ltree.grid(row=0, column=0, columnspan=10, padx=5, pady=5)
lysb.grid(row=0, column=10, sticky='ns')
# lxsb.grid(row=20, column=0, sticky='ew')
ltree.bind('<<TreeviewSelect>>', OnSelectL)


lechoLabel = Label(root, text='0', font=bft, width=5)
lechoLabel.grid(row=1, column=0, )
lserverTitle_Entry = Entry(root, textvariable=lserverTitleText, font=bft, width=35)
lserverTitle_Entry.grid(row=1, column=2, )
lswitchTitle_Entry = Entry(root, textvariable=lswitchTitleText, font=bft, width=35)
lswitchTitle_Entry.grid(row=1, column=4, )
laddButton = Button(root, text='添加线标', command=ladd, font=bft, takefocus=0)
laddButton.grid(row=1, column=5, )
limpButton = Button(root, text='线标导入', command=limp, font=bft, takefocus=0)
limpButton.grid(row=1, column=6,)
ldeleteButton = Button(root, text='删除选中', command=ldelete, font=bft, takefocus=0)
ldeleteButton.grid(row=1, column=7,)
leditButton = Button(root, text='修改选中', command=ledit, font=bft, takefocus=0)
leditButton.grid(row=1, column=8,)
lclearButton = Button(root, text='清空线标', command=lclear, font=bft, takefocus=0)
lclearButton.grid(row=1, column=9,)

lnoteLabel = Label(root, text='', font=bft, width=5)
lnoteLabel.grid(row=2, column=0, )
lserverInfo_Entry = Entry(root, textvariable=lserverInfoText, font=tft, width=40)
lserverInfo_Entry.grid(row=2, column=2, )
lswitchInfo_Entry = Entry(root, textvariable=lswitchInfoText, font=tft, width=40)
lswitchInfo_Entry.grid(row=2, column=4, )
lupButton = Button(root, text='关于...', command=about, font=bft, takefocus=0)
lupButton.grid(row=2, column=5,)
project_Entry = Entry(root, textvariable=projectText, font=bft, width=35)
project_Entry.grid(row=2, column=6, columnspan=4)


lmseparateline = Label(root, text='----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------')
lmseparateline.grid(row=3, columnspan=10)

frm = Frame(root)
frm.grid(row=4, column=0, columnspan=10)
mfrm = Frame(frm,)
# mfrm.grid(row=0, column=0, columnspan=5,)
mfrm.pack(side=LEFT)
mechoLabel = Label(mfrm, text='0', font=bft, width=5, pady=5)
mechoLabel.grid(row=0, column=0, )
mserverTitle_Entry = Entry(mfrm, textvariable=mserverTitleText, font=bft, width=25,)
mserverTitle_Entry.grid(row=0, column=1, columnspan=2)
mserverInfo_Entry = Entry(mfrm, textvariable=mserverInfoText, width=25)
mserverInfo_Entry.grid(row=0, column=3, columnspan=2)

maddButton = Button(mfrm, text='添加标签', command=madd, font=bft, takefocus=0,)
maddButton.grid(row=1, column=0, )
mimpButton = Button(mfrm, text='标签导入', command=mimp, font=bft, takefocus=0)
mimpButton.grid(row=1, column=1,)
mdeleteButton = Button(mfrm, text='删除选中', command=mdelete, font=bft, takefocus=0)
mdeleteButton.grid(row=1, column=2, )
meditButton = Button(mfrm, text='修改选中', command=medit, font=bft, takefocus=0)
meditButton.grid(row=1, column=3,)
mclearButton = Button(mfrm, text='清空标签', command=mclear, font=bft, takefocus=0)
mclearButton.grid(row=1, column=4,)

mtree = ttk.Treeview(mfrm, columns=('col1', 'col2'))
mtree.column('#0', width=10, anchor='w')
mtree.column('col1', width=240, anchor='center')
mtree.column('col2', width=240, anchor='center')
mtree.heading('col1', text='服务器标签头', anchor='center',)
mtree.heading('col2', text='服务器标签值', anchor='center')
mysb = ttk.Scrollbar(mfrm, orient='vertical', command=mtree.yview)
mxsb = ttk.Scrollbar(mfrm, orient='horizontal', command=mtree.xview)
mtree.configure(yscroll=mysb.set, xscroll=mxsb.set)
# mtree.heading('#0', text='Path', anchor='w')
mtree.grid(row=2, column=0, columnspan=5, padx=5, pady=5)
mysb.grid(row=2, column=5, sticky='ns')
mtree.bind('<<TreeviewSelect>>', OnSelectM)

efrm = Frame(frm,)
efrm.pack(side=RIGHT)
# efrm.grid(row=0, column=6, columnspan=5)
eechoLabel = Label(efrm, text='0', font=bft, width=5, pady=5)
eechoLabel.grid(row=0, column=0, )
eserverTitle_Entry = Entry(efrm, textvariable=eserverTitleText, font=bft, width=25)
eserverTitle_Entry.grid(row=0, column=1, columnspan=2)
eserverInfo_Entry = Entry(efrm, textvariable=eserverInfoText, width=25)
eserverInfo_Entry.grid(row=0, column=3, columnspan=2)

eaddButton = Button(efrm, text='添加标签', command=eadd, font=bft, takefocus=0)
eaddButton.grid(row=1, column=0, )
eimpButton = Button(efrm, text='标签导入', command=eimp, font=bft, takefocus=0)
eimpButton.grid(row=1, column=1,)
edeleteButton = Button(efrm, text='删除选中', command=edelete, font=bft, takefocus=0)
edeleteButton.grid(row=1, column=2, )
eeditButton = Button(efrm, text='修改选中', command=eedit, font=bft, takefocus=0)
eeditButton.grid(row=1, column=3,)
eclearButton = Button(efrm, text='清空标签', command=eclear, font=bft, takefocus=0)
eclearButton.grid(row=1, column=4,)

etree = ttk.Treeview(efrm, columns=('col1', 'col2'))
etree.column('#0', width=10, anchor='w')
etree.column('col1', width=240, anchor='center')
etree.column('col2', width=240, anchor='center')
etree.heading('col1', text='电源标签头', anchor='center',)
etree.heading('col2', text='电源标签值', anchor='center')
eysb = ttk.Scrollbar(efrm, orient='vertical', command=etree.yview)
exsb = ttk.Scrollbar(efrm, orient='horizontal', command=etree.xview)
etree.configure(yscroll=eysb.set, xscroll=exsb.set)
# etree.heading('#0', text='Path', anchor='w')
etree.grid(row=2, column=0, columnspan=5, padx=5, pady=5)
eysb.grid(row=2, column=5, sticky='ns')
etree.bind('<<TreeviewSelect>>', OnSelectE)

separateline = Label(root, text='*********************************************************************************************************************************************************************************************************')
separateline.grid(row=5, columnspan=10)

impButton = Button(root, text='全局导入', command=imp, font=bft, padx=3, pady=3)
impButton.grid(row=6, column=1, columnspan=2)
generateButton = Button(root, text='表格生成', command=generate, font=bft, padx=3, pady=3)
generateButton.grid(row=6, column=3, columnspan=2)
clearButton = Button(root, text='全部清空', command=clear, font=bft, padx=3, pady=3)
clearButton.grid(row=6, column=6, columnspan=2)

endline = Label(root, fg='green', text='欢迎使用', anchor='center', padx=0, pady=3)
endline.grid(row=7, column=1, columnspan=8,)

root.wm_title('哈尔滨银行标签生成器 V1.0')
root.iconbitmap('logo.ico')
root.mainloop()
