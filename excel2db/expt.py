from openpyxl import *
from openpyxl.utils import get_column_letter
import pymysql
import random
import re

trans = {'roomlocation': '机房', 'racknum': '机柜号', 'ulocation': '设备槽位', 'brand': '设备品牌', 'model': '设备型号',
         'type': '设备类型', 'mt': '设备MT号', 'sn': '设备SN号', 'project': '所属项目', 'describe': '设备用途',
         'uptime': '上架时间', 'belong': '所属部门', 'principal': '项目负责人', 'implementor': '实施负责人', 'mgmtip': '管理口IP', 'manageip': '管理IP', 'workip': '业务IP',
         'networknotes': '网络备注', 'system': '操作系统', 'software': '系统软件', 'hostname': '主机名', 'mgmtaccount': '管理口账号',
         'manageaccount': '管理账号', 'workaccount': '业务账号', 'accountnotes': '账号备注', 'psu': '电源单元', 'wiringinfo': '布线信息',
         'cpu': 'CPU', 'hdisk': '硬盘', 'memory': '内存', 'hardwarenotes': '硬件备注', 'cluster': '集群类型', 'ansible': '自动化运维', 'monitored': '监控状态',
         'inntp': '时钟同步', 'status': '设备状态', 'notes': '备注'
         }

config = {
    'host': '110.1.1.30',
    'port': 3306,
    'user': 'root',
    'passwd': 'Admin#233',
    'db': 'datacenter',
    'charset': 'utf8'
}


# create a connection
# conf: parameters of the connection
def connDB(conf):
    conn = pymysql.connect(**conf)
    return conn


# return query result
# conn: connection name;
# sql: query statement;
# args: query parameters;
# fetchnum: fetch number.
def exeQuery(conn, tsql, args, fetchnum):
    res = None
    try:
        with conn.cursor(pymysql.cursors.DictCursor) as cur:
            cur.execute(tsql, args)
            if fetchnum is None or fetchnum == -1:
                res = cur.fetchall()
            else:
                res = cur.fetchmany(fetchnum)
        conn.commit()
    finally:
        conn.close()
    return res


def exeUpdate(conn, tsql, args):
    res = None
    try:
        with conn.cursor() as cur:
            res = cur.execute(tsql, args)
        conn.commit()
    finally:
        conn.close()
    return res


def connClose(conn, cur):
    cur.close()
    conn.close()


# export data from Excel file
# filepath: path of target Excel file
# sheet: the sheet of Excel file need to be exported
# beginrow: begin exporting from this row
# begincol: begin exporting from this column
def expData(filepath, sheet, beginrow, begincol):
    title = None
    datalist = []
    coldict = {}
    #rowdict = {}
    colnum = 0
    rownum = beginrow - 1
    wb = load_workbook(filepath)
    ws = wb[sheet]

    # get the real count of columns
    for colindex in range(1, ws.max_column):
        #print(chr(ord('A') + colindex) + str(beginrow-1))
        if ws[get_column_letter(colindex) + str(beginrow-1)].value is not None:
            colnum += 1
            #coldict[get_column_letter(colindex)] = ws[get_column_letter(colindex) + str(beginrow-1)].value.strip()
        else:
            break

    # get the real count of rows
    for rowindex in range(beginrow, ws.max_row + 1):
        if ws['A' + str(rowindex)].value is not None or ws['B' + str(rowindex)].value is not None or ws['C' + str(rowindex)].value is not None or ws['D' + str(rowindex)].value is not None:
            rownum += 1
        else:
            break

    # export data
    for row in range(beginrow, rownum + 1):
        if ws['A' + str(row)].value is not None and '-' in str(ws['A' + str(row)].value):
            title = ws['A' + str(row)].value.strip()
            if row - 1 >= beginrow and ws['A' + str(row-1)].value is not None and '-' in str(ws['A' + str(row-1)].value):
                datalist.append({'roomlocation': sheet.strip(), 'racknum': ws['A' + str(row-1)].value.strip(), 'ulocation': '0-0'})
            print(title)
            continue
        else:
            if title is not None:
                rowdata = {'roomlocation': sheet.strip()}
                rowdata['racknum'] = title
                for col in range(1, ws.max_column + 1):
                    #colname = coldict[chr(ord(begincol) + col - (ord(begincol) - ord('A')))]
                    colname = ws[get_column_letter(col) + str(beginrow-1)].value
                    colvalue = ws[get_column_letter(col) + str(row)].value
                    if colvalue is not None:
                        colvalue = str(colvalue).strip()
                    if colvalue is None or colvalue == '':
                        colvalue = '-'

                    if colname == '序号':
                        print(colvalue)
                        rowdata['racknum'] = title
                    elif colname == '设备品牌':
                        rowdata['brand'] = colvalue
                        if colvalue is not None and colvalue == 'HUAWEI':
                            rowdata['mgmtip'] = '192.168.2.100'
                            rowdata['mgmtaccount'] = 'root/Huawei12#$'
                        elif colvalue is not None and colvalue == 'Inspur':
                            rowdata['mgmtaccount'] = 'admin/admin(root/superuser albert/admin)'
                    elif colname == '设备型号':
                        rowdata['model'] = colvalue
                    elif colname == '设备类型':
                        rowdata['type'] = colvalue
                        if colvalue == '网络交换机' or colvalue == '防火墙' or colvalue == '漏洞扫描' or colvalue == '路由器' or colvalue == '数模转换' or colvalue == '动态口令' or colvalue == '数字签名' or colvalue == '网络设备':
                            rowdata['belong'] = '网络组'
                        else:
                            rowdata['belong'] = '系统组'
                        if colvalue == '小型机':
                            rowdata['mgmtip'] = '192.168.2.147 192.168.3.147'
                            rowdata['mgmtaccount'] = 'admin/admin'
                            rowdata['manageaccount'] = 'root/root'
                        if colvalue == 'HMC':
                            rowdata['manageaccount'] = 'root/passw0rd'
                            rowdata['workaccount'] = 'hscroot/abc123'
                    elif colname == '所属项目':
                        rowdata['project'] = colvalue
                    elif colname == '设备用途':
                        rowdata['describe'] = colvalue
                        rowdata['status'] = '在线'
                        if colvalue is not None:
                            if '下线' in colvalue:
                                rowdata['status'] = '下线'
                            if '下电' in colvalue:
                                rowdata['status'] = '下电'
                            if '关机' in colvalue:
                                rowdata['status'] = '关机'
                    elif colname == '应用管理员':
                        rowdata['principal'] = colvalue
                    elif colname == '系统管理员':
                        rowdata['implementor'] = colvalue
                    elif colname == '网络管理员':
                        if colvalue == '-':
                            rowdata['netmanager'] = '于乃平/13946019925'
                        else:
                            rowdata['netmanager'] = colvalue
                    elif colname == '备份管理员':
                        if colvalue == '-':
                            rowdata['backuper'] = '董立国/18945661336'
                        else:
                            rowdata['backuper'] = colvalue
                    elif colname == '上架时间':
                        if colvalue != '-':
                            colvalue = colvalue.replace('/', '.').replace('-', '.')
                            if colvalue.count('.') < 2:
                                if colvalue.count('.') < 1:
                                    colvalue = colvalue + '.' + str(random.randint(1, 12))
                                colvalue = colvalue + '.' + str(random.randint(1, 28))

                                (y, m, d) = colvalue.split('.')
                                if len(y) < 3:
                                    y = '20' + y
                                colvalue = '%04d.%02d.%02d' % (int(y), int(m), int(d))
                        rowdata['uptime'] = colvalue
                    elif colname == '入库时间':
                        if colvalue != '-':
                            colvalue = colvalue.replace('/', '.').replace('-', '.')
                            if colvalue.count('.') < 2:
                                if colvalue.count('.') < 1:
                                    colvalue = colvalue + '.' + str(random.randint(1, 12))
                                colvalue = colvalue + '.' + str(random.randint(1, 28))

                                (y, m, d) = colvalue.split('.')
                                if len(y) < 3:
                                    y = '20' + y
                                colvalue = '%04d.%02d.%02d' % (int(y), int(m), int(d))
                        rowdata['stocktime'] = colvalue
                    elif colname == '设备MT号':
                        rowdata['mt'] = colvalue
                    elif colname == '设备SN号':
                        rowdata['sn'] = colvalue
                    elif colname == '设备槽位':
                        rowdata['ulocation'] = colvalue
                    elif colname == '设备区域':
                        rowdata['area'] = colvalue
                    elif colname == '设备标签':
                        rowdata['tag'] = colvalue
                    elif colname == '管理口IP':
                        if colvalue is not None:
                            rowdata['mgmtip'] = colvalue
                    elif colname == '管理IP':
                        if colvalue is not None:
                            rowdata['manageip'] = colvalue
                    elif colname == '业务IP':
                        rowdata['workip'] = colvalue
                    elif colname == '网络备注':
                        rowdata['networknotes'] = colvalue
                    elif colname == '主机名称':
                        rowdata['hostname'] = colvalue
                    elif colname == '操作系统':
                        rowdata['system'] = colvalue
                    elif colname == '应用软件':
                        rowdata['software'] = colvalue
                    elif colname == '管理口账号':
                        if colvalue is not None:
                            if colvalue != '/':
                                rowdata['mgmtaccount'] = colvalue
                    elif colname == '管理账号':
                        if colvalue is not None:
                            if colvalue != '/':
                                rowdata['manageaccount'] = colvalue
                    elif colname == '业务账号':
                        if colvalue != '/':
                            rowdata['workaccount'] = colvalue
                    elif colname == '账号备注':
                        rowdata['accountnotes'] = colvalue
                    elif colname == '电源单元':
                        if colvalue is None or colvalue == '-':
                            if rowdata['ulocation'] is not None and rowdata['ulocation'] != '0-0' and rowdata['ulocation'] != '-' and rowdata['ulocation'] != '':
                                if '-' in rowdata['ulocation']:
                                    umin, umax = rowdata['ulocation'].split('-')
                                    rowdata['psu'] = str(int(umax) - int(umin) + 1)
                                else:
                                    rowdata['psu'] = '2'
                        else:
                            rowdata['psu'] = colvalue
                    elif colname == '登录方式':
                        rowdata['loginmethod'] = colvalue
                    elif colname == '集群类型':
                        rowdata['cluster'] = colvalue
                    elif colname == '备份类型':
                        rowdata['backup'] = colvalue
                    elif colname == '资产编码':
                        rowdata['assetnum'] = colvalue
                    elif colname == '采购立项':
                        rowdata['approve'] = colvalue
                    elif colname == '所属银行':
                        rowdata['domain'] = colvalue
                    elif colname == '所属部门':
                        rowdata['belong'] = colvalue
                    elif colname == '设备状态':
                        if colvalue is None or colvalue == '-':
                            rowdata['status'] = '未知'
                        else:
                            rowdata['status'] = colvalue
                    elif colname == '监控状态':
                        if colvalue is None or colvalue == '-':
                            rowdata['monitored'] = '未知'
                        elif colvalue == '是' or colvalue.upper() == 'YES' or colvalue.upper() == 'Y':
                            rowdata['monitored'] = '开启'
                        else:
                            rowdata['monitored'] = '关闭'
                    elif colname == '时钟同步':
                        if colvalue is None or colvalue == '-':
                            rowdata['inntp'] = '未知'
                        elif colvalue == '是' or colvalue.upper() == 'YES' or colvalue.upper() == 'Y':
                            rowdata['inntp'] = '开启'
                        else:
                            rowdata['inntp'] = '关闭'
                    elif colname == '运维自动化':
                        if colvalue is None or colvalue == '-':
                            rowdata['ansible'] = '未知'
                        elif colvalue == '是' or colvalue.upper() == 'YES' or colvalue.upper() == 'Y':
                            rowdata['ansible'] = '开启'
                        else:
                            rowdata['ansible'] = '关闭'
                    elif colname == '资源':
                        cpuinfo = None
                        meminfo = None
                        diskinfo = None
                        if colvalue is not None and colvalue.replace(' ', '') != '':
                            colvalue = colvalue.replace('：', ':').replace('（', '(').replace('）', ')').replace('CPU', 'cpu').replace('MEM', 'mem').replace('DISK', 'disk').replace('storage', 'disk').replace('STORAGE', 'disk')
                            srcinfo = colvalue.split(' ')
                            for info in srcinfo:
                                if 'cpu' in info:
                                    cpuinfo = info.replace('cpu:', '').strip()
                                elif 'mem' in info:
                                    meminfo = info.replace('mem:', '').strip()
                                elif 'disk' in info:
                                    diskinfo = info.replace('disk:', '').strip().replace('(', '').replace(')', '')
                        rowdata['cpu'] = cpuinfo
                        rowdata['memory'] = meminfo
                        rowdata['hdisk'] = diskinfo

                    elif colname == 'CPU配置':
                        rowdata['cpu'] = colvalue
                    elif colname == '内存配置':
                        rowdata['memory'] = colvalue
                    elif colname == '硬盘配置':
                        rowdata['hdisk'] = colvalue
                    elif colname == 'PCI配置':
                        rowdata['pci'] = colvalue

                    elif colname == '入保时间':
                        if colvalue != '-':
                            colvalue = colvalue.replace('/', '.').replace('-', '.').replace('00:00:00', '').strip()
                            if colvalue.count('.') < 2:
                                if colvalue.count('.') < 1:
                                    colvalue = colvalue + '.' + str(random.randint(1, 12))
                                colvalue = colvalue + '.' + str(random.randint(1, 28))

                                print(colvalue)
                                (y, m, d) = colvalue.split('.')
                                if len(y) < 3:
                                    y = '20' + y
                                colvalue = '%04d.%02d.%02d' % (int(y), int(m), int(d))
                                print(colvalue)
                        rowdata['inmttime'] = colvalue

                    elif colname == '出保时间':
                        if colvalue != '-':
                            colvalue = colvalue.replace('/', '.').replace('-', '.').replace('00:00:00', '').strip()
                            if colvalue.count('.') < 2:
                                if colvalue.count('.') < 1:
                                    colvalue = colvalue + '.' + str(random.randint(1, 12))
                                colvalue = colvalue + '.' + str(random.randint(1, 28))

                                (y, m, d) = colvalue.split('.')
                                if len(y) < 3:
                                    y = '20' + y
                                colvalue = '%04d.%02d.%02d' % (int(y), int(m), int(d))
                        rowdata['outmttime'] = colvalue
                    elif colname == '供货厂商':
                        rowdata['spvendor'] = colvalue
                    elif colname == '集成厂商':
                        rowdata['itvendor'] = colvalue
                    elif colname == '维保厂商':
                        if colvalue is not None:
                            if '维保' in colvalue or '续保' in colvalue:
                                rowdata['mtnotes'] = colvalue
                            else:
                                rowdata['mtvendor'] = colvalue
                    elif colname == '维保备注':
                        rowdata['mtnotes'] = colvalue

                    elif colname == '备注':
                        rowdata['notes'] = colvalue
                        if colvalue is not None:
                            if '下线' in colvalue:
                                rowdata['status'] = '下线'
                            if '下电' in colvalue:
                                rowdata['status'] = '下电'
                            if '关机' in colvalue:
                                rowdata['status'] = '关机'
                    if '库房' in rowdata['roomlocation']:
                        rowdata['status'] = '库存'
                rowdata['hardwarenotes'] = '-'
                datalist.append(rowdata)

                    #print('     ' + colname + ':' + str(colvalue), end='')
                #print('\n')
    return datalist


def impData(data):
    res = ''
    connection = connDB(config)
    for item in data:
        colname = ''
        colloc = ''
        colvalue = []
        for col in item:
            if item[col] is not None:
                colname += '`' + col + '`, '
                colloc += '%s, '
                colvalue.append(item[col])
        colname = colname[:-2]
        colloc = colloc[:-2]
        tsql = 'insert ignore into device(' + colname + ') values(' + colloc + ')'
        try:
            with connection.cursor() as cur:
                res = cur.execute(tsql, colvalue)
        except Exception as ex:
            print(str(ex))
            #print(tsql)
    connection.commit()
    connection.close()
    print(res)
        #sq = "select roomlocation, racknum from device"
        #sql = 'insert into device(roomlocation, racknum, ulocation) values(%s, %s, %s)'
        #params = ('国裕三楼', 'RA-01', '9-12')
        #result = exeQuery(connection, sq, None, None)
        #columns = ','.join(item.keys())
        #values = item.values()
        #for ckey in item:
            #print(ckey)
'''

dds = expData('a.xlsx', '应用机房', 5, 'B')
for item in dds:
    print(str(len(item))+':'+item['roomlocation']+':'+item['racknum']+':'+item['brand']+':'+item['sn']+':'+item['netmanager'])
print(len(dds))
'''
impData(expData('a.xlsx', '应用机房', 5, 'B'))
impData(expData('a.xlsx', '核心机房', 5, 'B'))
#impData(expData('a.xlsx', '国裕三楼', 5, 'B'))


# connection.commit()
# connClose(connection, cursor)
